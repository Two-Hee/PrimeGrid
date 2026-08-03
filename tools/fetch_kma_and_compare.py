"""기상청 API 허브에서 일조·일사 자료를 받아 워크북 원자료와 대조한다.

이 저장소가 실행되는 원격 환경은 기상청 도메인이 네트워크 정책으로 차단되어 있으므로,
네트워크가 열린 PC에서 실행하는 것을 전제로 한다.

준비:
    export KMA_API_KEY="발급받은 인증키"      # 키를 소스나 인자에 직접 적지 말 것
    pip install openpyxl

사용:
    # 1) 수집 + 대조 (원본 응답은 --cache 디렉터리에 보관)
    python tools/fetch_kma_and_compare.py \
        --workbook 태양광_발전량_성능진단_기준모델_v1.0.xlsx \
        --cache kma_raw --report kma_compare.csv

    # 2) 이미 받아 둔 응답으로 대조만 다시 실행 (네트워크 불필요)
    python tools/fetch_kma_and_compare.py --workbook <xlsx> --cache kma_raw --offline

주의:
    - 권역별 대표 관측소 1곳만 조회한다. 워크북의 '권역' 값이 여러 관측소의 평균이라면
      당연히 차이가 난다. 그 경우 차이의 크기·부호 패턴이 평균 구성의 단서가 된다.
    - 일사(SI) 관측을 하지 않는 지점은 결측으로 보고되며, 리포트에 그대로 표시된다.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict

import openpyxl

API_URL = "https://apihub.kma.go.kr/api/typ01/url/sun_sfc_day.php"

# 워크북 '14_지역·상수' 시트의 권역 대표 도시에 대응하는 기상청 지점번호
STATIONS = {
    "서울경기": (108, "서울"),
    "강원영동": (105, "강릉"),
    "강원영서": (101, "춘천"),
    "충북": (131, "청주"),
    "충남": (133, "대전"),
    "경북": (143, "대구"),
    "경남": (159, "부산"),
    "전북": (146, "전주"),
    "전남": (156, "광주"),
    "제주": (184, "제주"),
}

MJ_PER_KWH = 3.6
MISSING = {"-9", "-99", "-999", "-9.0", "-99.0", "-999.0", "", "-"}


# ---------------------------------------------------------------- 수집
def cache_path(cache_dir: str, stn: int, year: int) -> str:
    return os.path.join(cache_dir, f"sun_sfc_day_{stn}_{year}.txt")


def fetch_year(stn: int, year: int, key: str, cache_dir: str, offline: bool) -> str:
    path = cache_path(cache_dir, stn, year)
    if os.path.exists(path):
        with open(path, encoding="utf-8", errors="replace") as fh:
            return fh.read()
    if offline:
        raise FileNotFoundError(f"캐시 없음: {path} (오프라인 모드)")

    query = urllib.parse.urlencode({
        "authKey": key, "stn": stn, "tm1": f"{year}0101", "tm2": f"{year}1231",
    })
    with urllib.request.urlopen(f"{API_URL}?{query}", timeout=60) as resp:
        text = resp.read().decode("utf-8", errors="replace")
    os.makedirs(cache_dir, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text)
    time.sleep(0.5)   # 서버 부하 완화
    return text


# ---------------------------------------------------------------- 파싱
def find_header(lines) -> list[str] | None:
    """'#' 주석 중 컬럼명이 나열된 줄을 찾아 컬럼 목록을 돌려준다."""
    for line in lines:
        if not line.startswith("#"):
            continue
        cols = line.lstrip("#").split()
        if "STN" in cols and any(c.startswith("SS") or c.startswith("SI") for c in cols):
            return cols
    return None


def pick(cols: list[str], *prefixes: str) -> int | None:
    for want in prefixes:
        for i, c in enumerate(cols):
            if c.upper() == want:
                return i
    for want in prefixes:
        for i, c in enumerate(cols):
            if c.upper().startswith(want):
                return i
    return None


def parse(text: str):
    """(YYYYMMDD, 일조시간, 일사량MJ) 목록과 사용한 컬럼명을 돌려준다."""
    lines = text.splitlines()
    cols = find_header(lines)
    if cols is None:
        raise ValueError("응답에서 컬럼 머리글을 찾지 못했습니다. 원본을 확인하세요.")
    i_tm = pick(cols, "TM", "YYMMDD")
    i_ss = pick(cols, "SS_DAY", "SS")
    i_si = pick(cols, "SI_DAY", "SI")
    if None in (i_tm, i_ss, i_si):
        raise ValueError(f"필요한 컬럼을 찾지 못했습니다. 머리글: {cols}")

    out = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        f = line.split()
        if len(f) <= max(i_tm, i_ss, i_si):
            continue
        tm = f[i_tm][:8]
        if not tm.isdigit():
            continue
        def num(tok):
            return None if tok in MISSING or float(tok) < 0 else float(tok)
        try:
            out.append((tm, num(f[i_ss]), num(f[i_si])))
        except ValueError:
            continue
    return out, (cols[i_tm], cols[i_ss], cols[i_si])


def monthly(daily):
    """일별 값을 월 합계로 집계한다. 결측일이 있으면 해당 월을 결측으로 처리한다."""
    ss_sum, si_sum, ss_missing, si_missing, days = (defaultdict(float), defaultdict(float),
                                                    defaultdict(int), defaultdict(int),
                                                    defaultdict(int))
    for tm, ss, si in daily:
        key = (int(tm[:4]), int(tm[4:6]))
        days[key] += 1
        if ss is None:
            ss_missing[key] += 1
        else:
            ss_sum[key] += ss
        if si is None:
            si_missing[key] += 1
        else:
            si_sum[key] += si
    res = {}
    for key in days:
        res[key] = {
            "일조합_hr": None if ss_missing[key] else round(ss_sum[key], 1),
            "일사합_MJ": None if si_missing[key] else si_sum[key],
            "일수": days[key],
            "결측일_일조": ss_missing[key],
            "결측일_일사": si_missing[key],
        }
    return res


# ---------------------------------------------------------------- 대조
def read_workbook(path: str):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["20_원자료"] if "20_원자료" in wb.sheetnames else wb["원자료"]
    data = {}
    for r in range(2, ws.max_row + 1):
        region = ws.cell(row=r, column=1).value
        if region is None:
            continue
        data[(region, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)] = {
            "일조합_hr": ws.cell(row=r, column=5).value,
            "일사합_kWh": ws.cell(row=r, column=6).value,
        }
    return data


def main() -> int:
    ap = argparse.ArgumentParser(description="기상청 일조·일사 자료와 워크북 원자료 대조")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--cache", default="kma_raw")
    ap.add_argument("--report", default="kma_compare.csv")
    ap.add_argument("--years", default="2016-2025")
    ap.add_argument("--offline", action="store_true", help="캐시만 사용(네트워크 미사용)")
    ap.add_argument("--tol", type=float, default=0.5, help="일치로 볼 상대오차 한계(%%)")
    args = ap.parse_args()

    y0, y1 = (int(x) for x in args.years.split("-"))
    key = os.environ.get("KMA_API_KEY", "")
    if not key and not args.offline:
        print("KMA_API_KEY 환경변수가 필요합니다. (--offline 이면 캐시만 사용)", file=sys.stderr)
        return 2

    book = read_workbook(args.workbook)
    rows, stats = [], defaultdict(int)
    header_note = None

    for region, (stn, city) in STATIONS.items():
        for year in range(y0, y1 + 1):
            try:
                text = fetch_year(stn, year, key, args.cache, args.offline)
                daily, used = parse(text)
            except Exception as exc:                      # 수집·파싱 실패는 행으로 남긴다
                stats["수집실패"] += 1
                rows.append([region, city, stn, year, "", "", "", "", "", "", f"오류: {exc}"])
                continue
            header_note = header_note or used
            for (yy, mm), agg in sorted(monthly(daily).items()):
                ref = book.get((region, yy, mm))
                si_kwh = None if agg["일사합_MJ"] is None else agg["일사합_MJ"] / MJ_PER_KWH
                if ref is None:
                    verdict, diff_pct = "워크북에 없음", ""
                elif si_kwh is None:
                    verdict, diff_pct = "관측값 결측", ""
                    stats["결측"] += 1
                else:
                    d = si_kwh - ref["일사합_kWh"]
                    diff_pct = 100 * d / ref["일사합_kWh"] if ref["일사합_kWh"] else 0
                    if abs(diff_pct) <= args.tol:
                        verdict = "일치"
                        stats["일치"] += 1
                    else:
                        verdict = "불일치"
                        stats["불일치"] += 1
                    diff_pct = round(diff_pct, 3)
                rows.append([
                    region, city, stn, yy, mm,
                    "" if si_kwh is None else round(si_kwh, 4),
                    "" if ref is None else ref["일사합_kWh"],
                    "" if agg["일조합_hr"] is None else agg["일조합_hr"],
                    "" if ref is None else ref["일조합_hr"],
                    agg["결측일_일사"], verdict if not diff_pct else f"{verdict} ({diff_pct}%)",
                ])

    with open(args.report, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["권역", "대표지점", "지점번호", "연도", "월",
                    "기상청_일사_kWh", "워크북_일사_kWh",
                    "기상청_일조_hr", "워크북_일조_hr", "일사결측일수", "판정"])
        w.writerows(rows)

    print(f"사용 컬럼: {header_note}")
    print(f"일치 {stats['일치']} / 불일치 {stats['불일치']} / 결측 {stats['결측']} / 수집실패 {stats['수집실패']}")
    print(f"리포트: {args.report}")
    if stats["불일치"]:
        print("→ 불일치가 계통적으로 나타나면 권역 값이 단일 지점이 아니라 여러 지점의 평균일 수 있습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
