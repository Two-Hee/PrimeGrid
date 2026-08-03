"""태양광 발전량 성능진단 기준모델 v1.0 워크북 생성 스크립트.

기존에 두 개의 문서(프로그램 기준모델 명세 / 일사량·발전량 계산 워크북)가
하나의 파일로 합쳐져 있던 자료를 정리하여 단일 체계의 워크북으로 재구성한다.

사용법:
    python tools/build_v1_workbook.py <원본.xlsx> <출력.xlsx>

원본에서 가져오는 것은 '원자료' 시트의 관측 실측값(지역·연월별 일조/일사)뿐이며,
나머지 계산 시트는 모두 수식으로 재생성한다.
"""

from __future__ import annotations

import sys

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# --------------------------------------------------------------------------
# 시트 이름 (참조는 반드시 quote() 를 통해 생성한다)
# --------------------------------------------------------------------------
COVER = "01_표지·목차"
VER = "02_버전이력"
OVERVIEW = "03_모델개요"
CALC = "04_계산식_명세"
FACT = "05_FACTOR_기준값"
DIAG = "06_진단규칙"
DICT = "07_데이터사전"
TEST = "08_검증CASE"
REPORT = "09_보고서구성"
DEV = "10_개발연계"
MODEL = "11_발전량_모델"
MFACT = "12_월별FACTOR"
SCEN = "13_시나리오비교"
REGION = "14_지역·상수"
GEO = "15_태양기하"
LIMIT = "16_적용한계·참고"
RAW = "20_원자료"
PM = "21_피벗_월별"
PY = "22_피벗_연도별"
PS = "23_피벗_계절패턴"
CHART = "24_차트"
HELP = "_시나리오계산"

DOC_VERSION = "v1.0"
MODEL_VERSION = "MVP-0.2.0"
DOC_DATE = "2026-08-03"

REGIONS = ["서울경기", "강원영동", "강원영서", "충북", "충남", "경북", "경남", "전북", "전남", "제주"]
YEARS = list(range(2016, 2026))
MONTH_DOY = [17, 47, 75, 105, 135, 162, 198, 228, 258, 288, 318, 344]
STEPS = 96  # 15분 간격 = 하루 96 스텝


def q(sheet: str) -> str:
    """수식에서 사용할 시트 접두사."""
    return f"'{sheet}'!"


# --------------------------------------------------------------------------
# 서식 정의
# --------------------------------------------------------------------------
FONT = "맑은 고딕"

NAVY = "1F3864"
BLUE = "2F5597"
LIGHT = "D9E2F3"
LIGHTER = "EDF2FA"
GRAY = "808080"
INPUT_FILL = "FFF2CC"
BORDER_COLOR = "BFBFBF"

F_TITLE = Font(name=FONT, size=16, bold=True, color="FFFFFF")
F_SUBTITLE = Font(name=FONT, size=9, color="595959", italic=True)
F_SECTION = Font(name=FONT, size=11, bold=True, color=NAVY)
F_HEAD = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT, size=9.5)
F_BODY_B = Font(name=FONT, size=9.5, bold=True)
F_INPUT = Font(name=FONT, size=9.5, bold=True, color="0000FF")
F_LINK = Font(name=FONT, size=9.5, color="008000")
F_NOTE = Font(name=FONT, size=9, color="595959")
F_KPI = Font(name=FONT, size=11, bold=True, color=NAVY)
F_HYPER = Font(name=FONT, size=9.5, color="0563C1", underline="single")

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HEAD = PatternFill("solid", fgColor=BLUE)
FILL_SECTION = PatternFill("solid", fgColor=LIGHT)
FILL_BAND = PatternFill("solid", fgColor=LIGHTER)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_FILL)
FILL_KPI = PatternFill("solid", fgColor="E2EFDA")

_side = Side(style="thin", color=BORDER_COLOR)
BOX = Border(left=_side, right=_side, top=_side, bottom=_side)

AL_L = Alignment(horizontal="left", vertical="center")
AL_LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center")
AL_CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_R = Alignment(horizontal="right", vertical="center")

NUM_INT = "#,##0"
NUM_1 = "#,##0.0"
NUM_2 = "#,##0.00"
NUM_3 = "0.000"
NUM_PCT = "0.0%"


def title_block(ws, title: str, subtitle: str, last_col: str = "J") -> None:
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws.merge_cells(f"A1:{last_col}1")
    for col in range(1, openpyxl.utils.column_index_from_string(last_col) + 1):
        ws.cell(row=1, column=col).fill = FILL_TITLE
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = F_SUBTITLE
        ws["A2"].alignment = AL_L
        ws.merge_cells(f"A2:{last_col}2")
        ws.row_dimensions[2].height = 18


def section(ws, row: int, text: str, last_col: str) -> None:
    ws.cell(row=row, column=1, value=text).font = F_SECTION
    ws.merge_cells(f"A{row}:{last_col}{row}")
    for col in range(1, openpyxl.utils.column_index_from_string(last_col) + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION
    ws.row_dimensions[row].height = 20


def header_row(ws, row: int, headers, start_col: int = 1, height: int = 30) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_CW
        c.border = BOX
    ws.row_dimensions[row].height = height


def table_body(ws, first_row: int, rows, start_col: int = 1, wrap_cols=(), band: bool = True) -> None:
    """2차원 리스트를 표 본문으로 기록한다."""
    for r_off, data in enumerate(rows):
        r = first_row + r_off
        for c_off, v in enumerate(data):
            c = ws.cell(row=r, column=start_col + c_off, value=v)
            c.font = F_BODY
            c.border = BOX
            c.alignment = AL_LW if (c_off in wrap_cols) else AL_LW
            if band and r_off % 2 == 1:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 30


def widths(ws, spec: dict) -> None:
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def input_cell(ws, addr: str, value, fmt: str | None = None) -> None:
    c = ws[addr]
    c.value = value
    c.font = F_INPUT
    c.fill = FILL_INPUT
    c.border = BOX
    c.alignment = AL_C
    if fmt:
        c.number_format = fmt


def page_setup(ws, landscape: bool = True, rows_to_repeat: str | None = None) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if rows_to_repeat:
        ws.print_title_rows = rows_to_repeat


# ==========================================================================
# 01_표지·목차
# ==========================================================================
INDEX_ROWS = [
    ("01", COVER, "문서 개요, 요약 지표, 시트 목차", "표지"),
    ("02", VER, "모델·문서 버전 이력 및 구버전 시트 매핑", "기준문서"),
    ("03", OVERVIEW, "간편·상세·전문가 모드별 기능과 구현 범위", "기준문서"),
    ("04", CALC, "월별 계산 순서와 계산식 정의(코드·Excel 공통)", "기준문서"),
    ("05", FACT, "손실 FACTOR 기본값·권장범위·적용 기준", "기준문서"),
    ("06", DIAG, "저하원인 진단규칙과 임계값", "기준문서"),
    ("07", DICT, "입력·출력 데이터 사전(API/보고서 공통)", "기준문서"),
    ("08", TEST, "계산·진단 검증 CASE 및 허용오차", "기준문서"),
    ("09", REPORT, "PDF·PPT·Excel 산출물 구성", "기준문서"),
    ("10", DEV, "웹 프로그램 연계 컴포넌트와 검증 방법", "기준문서"),
    ("11", MODEL, "지역·경사·방위·PR 입력 → 월/연간 POA 및 발전량 계산", "계산모델"),
    ("12", MFACT, "월별 H0·Kt·산란비·Rb·View Factor 확인", "계산모델"),
    ("13", SCEN, "경사각·방위각 시나리오 10종 비교", "계산모델"),
    ("14", REGION, "지역 대표 위도와 모델 상수·기본값", "계산모델"),
    ("15", GEO, "15분 태양벡터 적분에 의한 Rb·H0 산출 근거", "계산모델"),
    ("16", LIMIT, "적용 한계, 입력 기준, 참고 문헌", "계산모델"),
    ("20", RAW, "기상청 지역별 월 일사량 원자료 1,200행(kWh 환산)", "데이터"),
    ("21", PM, "연월×지역 피벗(월 합계 / 일평균)", "데이터"),
    ("22", PY, "연도×지역 피벗(연 합계 / 일평균)", "데이터"),
    ("23", PS, "달력월 기후평균 피벗(월 합계 / 일평균)", "데이터"),
    ("24", CHART, "연도별 추이·계절 패턴 그래프", "데이터"),
]


def build_cover(wb):
    ws = wb.create_sheet(COVER)
    ws.sheet_view.showGridLines = False
    title_block(
        ws,
        "태양광 발전량 성능진단 기준모델",
        "지역별 일사량 → 경사면 일사량(POA) → 발전량 → 성능진단으로 이어지는 계산 기준을 정의하는 통합 워크북입니다.",
        "F",
    )

    meta = [
        ("문서 버전", DOC_VERSION, "본 워크북 정리 버전"),
        ("모델 버전", MODEL_VERSION, "웹 프로그램(app.py)과 공유하는 계산모델 버전"),
        ("작성 기준일", DOC_DATE, "버전이력 시트 참조"),
        ("분석 기간", "2016년 1월 ~ 2025년 12월 (10개년)", "연 단위 완전 데이터"),
        ("대상 지역", "10개 권역 (서울경기·강원영동·강원영서·충북·충남·경북·경남·전북·전남·제주)", ""),
        ("자료 출처", "기상청 기상자료개방포털 「기후통계분석 - 일조일사」 (data.kma.go.kr)", "MJ/m² ÷ 3.6 = kWh/m²"),
        ("용도", "예비 타당성·성능진단용 기준모델", "금융·성능보증용은 시간별 현장자료 적용 필요"),
    ]
    section(ws, 4, "■ 문서 정보", "F")
    r = 5
    for label, value, note in meta:
        ws.cell(row=r, column=1, value=label).font = F_BODY_B
        ws.cell(row=r, column=1).alignment = AL_L
        ws.cell(row=r, column=1).border = BOX
        c = ws.cell(row=r, column=2, value=value)
        c.font = F_BODY
        c.alignment = AL_LW
        c.border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col in range(2, 5):
            ws.cell(row=r, column=col).border = BOX
        n = ws.cell(row=r, column=5, value=note)
        n.font = F_NOTE
        n.alignment = AL_LW
        n.border = BOX
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.cell(row=r, column=6).border = BOX
        ws.row_dimensions[r].height = 20
        r += 1

    # 요약 지표
    section(ws, 13, "■ 요약 지표 (2016~2025 실측 자료 기준, 자동 계산)", "F")
    kpi = [
        ("연평균 일사량 최고 권역",
         f"=INDEX({q(PY)}$B$3:$K$3,MATCH(MAX({q(PY)}$B$14:$K$14),{q(PY)}$B$14:$K$14,0))",
         f"=MAX({q(PY)}$B$14:$K$14)", "kWh/m²/년"),
        ("일평균 일사량 최고 권역",
         f"=INDEX({q(PY)}$O$3:$X$3,MATCH(MAX({q(PY)}$O$14:$X$14),{q(PY)}$O$14:$X$14,0))",
         f"=MAX({q(PY)}$O$14:$X$14)", "kWh/m²/일"),
        ("전국 평균 일사량",
         "10개 권역 평균", f"=AVERAGE({q(PY)}$L$4:$L$13)", "kWh/m²/년"),
        ("일사량 최대 월(계절 패턴)",
         f"=INDEX({q(PS)}$A$4:$A$15,MATCH(MAX({q(PS)}$L$4:$L$15),{q(PS)}$L$4:$L$15,0))&\"월\"",
         f"=MAX({q(PS)}$L$4:$L$15)", "kWh/m²/월"),
        ("일사량 최소 월(계절 패턴)",
         f"=INDEX({q(PS)}$A$4:$A$15,MATCH(MIN({q(PS)}$L$4:$L$15),{q(PS)}$L$4:$L$15,0))&\"월\"",
         f"=MIN({q(PS)}$L$4:$L$15)", "kWh/m²/월"),
        ("현재 입력 기준 연간 발전량",
         f"={q(MODEL)}$B$5&\" / 경사 \"&{q(MODEL)}$B$7&\"° / 방위 \"&{q(MODEL)}$B$8&\"°\"",
         f"={q(MODEL)}$F$9", "MWh/년"),
    ]
    r = 14
    for label, sub, val, unit in kpi:
        ws.cell(row=r, column=1, value=label).font = F_BODY_B
        ws.cell(row=r, column=1).border = BOX
        ws.cell(row=r, column=1).alignment = AL_L
        c = ws.cell(row=r, column=2, value=sub)
        c.font = F_LINK
        c.alignment = AL_L
        c.border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=3).border = BOX
        v = ws.cell(row=r, column=4, value=val)
        v.font = F_KPI
        v.fill = FILL_KPI
        v.alignment = AL_R
        v.border = BOX
        v.number_format = NUM_2 if "일평균" in label else NUM_1
        u = ws.cell(row=r, column=5, value=unit)
        u.font = F_NOTE
        u.alignment = AL_L
        u.border = BOX
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.cell(row=r, column=6).border = BOX
        ws.row_dimensions[r].height = 20
        r += 1

    # 목차
    section(ws, 21, "■ 시트 목차", "F")
    header_row(ws, 22, ["No.", "시트", "내용", "구분"], height=22)
    ws.merge_cells("C22:E22")
    ws.cell(row=22, column=4).fill = FILL_HEAD
    ws.cell(row=22, column=5).fill = FILL_HEAD
    for i, (no, name, desc, kind) in enumerate(INDEX_ROWS):
        r = 23 + i
        ws.cell(row=r, column=1, value=no).font = F_BODY
        ws.cell(row=r, column=1).alignment = AL_C
        link = ws.cell(row=r, column=2, value=name)
        link.font = F_HYPER
        link.alignment = AL_L
        link.hyperlink = f"#'{name}'!A1"
        d = ws.cell(row=r, column=3, value=desc)
        d.font = F_BODY
        d.alignment = AL_LW
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        k = ws.cell(row=r, column=6, value=kind)
        k.font = F_BODY
        k.alignment = AL_C
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BOX
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 18

    note_r = 23 + len(INDEX_ROWS) + 1
    ws.cell(row=note_r, column=1,
            value="※ 노란 배경·파란 글씨 셀만 사용자가 직접 입력하는 값이며, 그 외 셀은 수식으로 계산됩니다. "
                  "초록 글씨는 다른 시트를 참조하는 값입니다.").font = F_NOTE
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)
    ws.cell(row=note_r, column=1).alignment = AL_LW

    widths(ws, {"A": 30, "B": 24, "C": 26, "D": 16, "E": 14, "F": 12})
    page_setup(ws, landscape=False)
    ws.sheet_properties.tabColor = "C00000"
    return ws


# ==========================================================================
# 02_버전이력
# ==========================================================================
VERSION_ROWS = [
    ["MVP-0.1.0", "2026-08-03", "최초 MVP", "간편·상세·전문가 모드, 월간 POA, PDF/PPT/Excel 산출", "완료", "완료"],
    ["MVP-0.2.0", "2026-08-03", "기능·품질 개선",
     "부분연도 동일기간 비교, 결측월 null 처리, SQLite 프로젝트 저장, 전문가 이미지 부록, PptxGenJS", "완료", "완료"],
    ["NEXT", "미정", "후속", "시간단위 pvlib 엔진, 주소/지도 입력, 유사 발전소 벤치마크, Windows 단일 실행형", "미착수", "미착수"],
]

SHEET_MAP = [
    ["버전관리", VER, "v1.0 이력 및 시트 매핑 추가"],
    ["프로그램_개요", OVERVIEW, "그대로 유지"],
    ["계산식_명세", CALC, "계산_가이드의 '모델 흐름' 통합"],
    ["FACTOR_정의", FACT, "지역_설정에 중복 정의되던 기본값의 단일 출처로 지정"],
    ["진단규칙_명세", DIAG, "그대로 유지"],
    ["데이터_사전", DICT, "그대로 유지"],
    ["검증_CASE", TEST, "중복된 CASE ID(TC-012 2건) 정정, 일련번호 재부여"],
    ["보고서_구성", REPORT, "그대로 유지"],
    ["개발_연계", DEV, "그대로 유지"],
    ["표지", COVER, "문서 맨 앞으로 이동, 목차 하이퍼링크·요약지표 추가"],
    ["발전량_모델", MODEL, "입력/결과 영역 구분, 입력 셀 표시 및 유효성 목록 추가"],
    ["FACTOR_월별", MFACT, "그대로 유지"],
    ["방위경사_비교", SCEN, "본표 옆에 있던 보조 계산 블록을 별도 숨김 시트로 분리"],
    ["지역_설정", REGION, "기본값 3종(PR·가동률·반사율)은 05 시트를 참조하도록 변경"],
    ["계산_가이드", f"{CALC} / {LIMIT}", "계산식은 04로, 한계·참고자료는 16으로 분리"],
    ["태양기하_상세", GEO, "15분 적분 상세행을 그룹으로 접어 요약부만 표시"],
    ["원자료", RAW, "그대로 유지(표 이름 RawDataKWh 유지)"],
    ["월별_일사량_kWh + 월별_일평균_kWh", PM, "월 합계·일평균 2개 시트를 1개로 통합"],
    ["연도별_일사량_kWh + 연도별_일평균_kWh", PY, "연 합계·일평균 통합, 10개년 평균 행 추가"],
    ["월평균_계절패턴_kWh + …_일평균_kWh", PS, "계절 패턴 2개 시트를 1개로 통합"],
    ["차트", CHART, "통합 피벗을 참조하도록 재작성"],
]


def build_version(wb):
    ws = wb.create_sheet(VER)
    title_block(ws, "버전 이력", "모델 버전과 문서 버전을 함께 관리하며, 계산식이 바뀌면 반드시 버전을 함께 갱신합니다.", "F")
    section(ws, 4, "■ 버전 이력", "F")
    header_row(ws, 5, ["모델 버전", "일자", "구분", "변경 내용", "검증 상태", "프로그램 반영"])
    rows = [["문서 " + DOC_VERSION, DOC_DATE, "자료 정리",
             "AI 생성 자료 2건을 단일 워크북으로 재구성(시트 순서·명명·서식 통일, 중복 정의 제거, 보조 계산 분리)",
             "완료", "해당 없음"]] + VERSION_ROWS
    table_body(ws, 6, rows)

    start = 6 + len(rows) + 1
    section(ws, start, "■ 구버전 → v1.0 시트 매핑", "F")
    header_row(ws, start + 1, ["구버전 시트", "v1.0 시트", "정리 내용"])
    ws.merge_cells(start_row=start + 1, start_column=3, end_row=start + 1, end_column=6)
    for col in range(3, 7):
        ws.cell(row=start + 1, column=col).fill = FILL_HEAD
        ws.cell(row=start + 1, column=col).border = BOX
    for i, (old, new, memo) in enumerate(SHEET_MAP):
        r = start + 2 + i
        for col, v in ((1, old), (2, new), (3, memo)):
            c = ws.cell(row=r, column=col, value=v)
            c.font = F_BODY
            c.alignment = AL_LW
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BOX
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 18

    widths(ws, {"A": 30, "B": 22, "C": 16, "D": 46, "E": 12, "F": 14})
    ws.freeze_panes = "A6"
    page_setup(ws)
    return ws


# ==========================================================================
# 03_모델개요
# ==========================================================================
OVERVIEW_ROWS = [
    ["주요 목적", "지역 장기평균 및 분석연도 기상 대비 실제 발전량 비교", "현장 일사량·설비·인버터 정보를 추가해 원인 후보 검토",
     "시간자료·항목별 손실·전문가 의견 및 보고서 편집", "구현", "유사 발전소 벤치마크"],
    ["최소 입력", "지역, 분석연도, DC/AC 용량, 실제 월별 발전량", "간편 입력 + 경사/방위, 모듈, 일사량, 인버터별 연간자료",
     "상세 입력 + 시간자료, MPPT, 상태코드, 전문가 손실 FACTOR", "구현", "시간자료 상세 진단"],
    ["기상 기준", "2016~2025 지역별 GHI 평균 및 선택연도 GHI", "현장 GHI 또는 POA 우선",
     "현장 POA·시간별 GHI/DNI/DHI 및 전사모델 선택", "월 단위 구현", "pvlib 시간엔진"],
    ["예상 발전량", "PR 기반", "PR 기반 또는 현장자료 보정", "항목별 손실 기반 선택", "구현", "모듈/인버터 사양 DB"],
    ["진단 결과", "정상/관찰/주의/점검권고, 기본 원인 후보", "인버터 편차·출력제한·클리핑 가능성 추가",
     "진단규칙 임계값과 전문가 의견 관리", "기본 규칙 구현", "이상감지 학습모델"],
    ["다운로드", "고객 PDF, PPT, 상세 Excel, JSON", "동일", "전문가 의견·현장사진 추가용 PPT",
     "구현(전문가 이미지 부록 포함)", "슬라이드 선택·브랜드 템플릿"],
    ["데이터 등급", "D 또는 C", "C 또는 B", "B 또는 A", "구현", "자동 품질점수 고도화"],
    ["권한", "고객", "고객/분석자", "전문가 PIN 및 서버 토큰", "서명 토큰·8시간 만료 구현", "회원·조직·RBAC"],
    ["로컬 저장", "브라우저 입력값 임시보관", "로컬 SQLite 프로젝트 저장·불러오기", "전문가 첨부파일 로컬 저장",
     "구현", "백업·복구/다중 사용자"],
]


def build_overview(wb):
    ws = wb.create_sheet(OVERVIEW)
    title_block(ws, "모델 개요 및 모드별 기능", "웹 프로그램, 상세 계산 Excel, 고객 PDF/PPT가 공유해야 하는 분석 원칙과 모드별 기능을 정의합니다.", "F")
    section(ws, 4, "■ 모드별 기능 정의", "F")
    header_row(ws, 5, ["구분", "간편 모드", "상세 모드", "전문가 모드", "구현 상태(MVP)", "후속 개발"])
    table_body(ws, 6, OVERVIEW_ROWS)
    widths(ws, {"A": 14, "B": 34, "C": 34, "D": 34, "E": 20, "F": 22})
    ws.freeze_panes = "B6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 04_계산식_명세
# ==========================================================================
FLOW_ROWS = [
    ["①", "수평면 일사량 선택", "지역·연도별 월 합계 GHI(원자료 또는 현장 실측)"],
    ["②", "월 일평균 환산", "H = 월 GHI ÷ 해당 월 일수"],
    ["③", "태양기하 계산", "대표 위도·월 대표일로 적위 δ, 일몰시간각 ωs, 대기권외 일사량 H0"],
    ["④", "확산 분리", "Kt = H/H0 → Erbs 월평균 상관식으로 산란비 Hd/H 추정"],
    ["⑤", "직달 전사", "15분 간격 태양벡터 적분으로 Rb 계산"],
    ["⑥", "경사면 일사량", "POA = 직달 + 천공산란 + 지면반사"],
    ["⑦", "발전량 산정", "POA × DC용량 × PR × 가동률 × (1-출력제한) × 기타보정"],
]

CALC_ROWS = [
    [1, "대표일", "representative_day", "1~12월 대표일: 17, 47, 75, 105, 135, 162, 198, 228, 258, 288, 318, 344",
     "Julian day", "전체", "월", "월평균 모델용"],
    [2, "태양 적위", "delta", "δ = 23.45 × sin[360 × (284+n)/365]", "degree", "전체", "대표일", "각도 계산 시 radian 변환"],
    [3, "일몰시간각", "omega_s", "ωs = acos(-tanφ × tanδ)", "radian", "전체", "위도 φ", "acos 인수 -1~1 제한"],
    [4, "대기권외 일사량", "h0_daily", "H0 = (24×60/π)×Gsc×dr×[ωs sinφ sinδ + cosφ cosδ sinωs] ÷ 3.6",
     "kWh/m²/day", "전체", "Gsc=0.082 MJ/m²/min", "0 이하 방지"],
    [5, "수평면 일평균", "h_daily", "H = 월 GHI ÷ 월 일수", "kWh/m²/day", "전체", "지역/현장 GHI", "결측을 0으로 대체하지 않음"],
    [6, "청명도지수", "kt_raw / kt_applied", "Kt_raw = H/H0; Erbs 적용 Kt = MAX(0.30, MIN(0.80, Kt_raw))",
     "-", "전체", "H, H0", "원자료 Kt와 적용 Kt를 모두 저장"],
    [7, "산란비", "diffuse_fraction",
     "ωs ≤ 81.4°: 1.391-3.560Kt+4.189Kt²-2.137Kt³ / 그 외: 1.311-3.022Kt+3.427Kt²-1.821Kt³",
     "-", "전체", "Erbs 월평균식", "0.05~1.00 제한"],
    [8, "직달·산란 분리", "beam / diffuse", "Hd = H×Kd, Hb = MAX(0, H-Hd)", "kWh/m²/day", "전체", "H, Kd", "월평균 분리값"],
    [9, "직달 경사계수", "rb", "15분 간격 태양벡터 적분: Rb = Σmax(cosθ,0) / Σmax(cosθz,0)",
     "-", "전체", "위도, 경사각 β, 방위각 A", "방위각 북0 동90 남180 서270"],
    [10, "천공 View Factor", "sky_factor", "(1+cosβ)/2", "-", "전체", "경사각", "등방성 천공 모델"],
    [11, "지면 View Factor", "ground_factor", "(1-cosβ)/2", "-", "전체", "경사각", "지면반사율 별도 적용"],
    [12, "경사면 일사량", "poa_month", "POA = [Hb×Rb + Hd×sky_factor + H×ρg×ground_factor] × 일수",
     "kWh/m²/month", "전체", "GHI, 경사, 방위, 반사율", "현장 POA 입력 시 계산값보다 우선"],
    [13, "모듈 경년열화", "age_factor", "F_age = (1 - 연간열화율)^운영연수", "-", "전체", "상업운전일, 분석연도",
     "상업운전일 미입력 시 1.000"],
    [14, "PR 기반 예상량", "expected_pr",
     "E = POA × Pdc × PR × F_age × Availability × (1-Curtailment) × F_other", "kWh", "간편/상세",
     "PR, 가동률, 출력제한", "PR에 포함된 손실 중복 차감 금지"],
    [15, "항목별 예상량", "expected_component",
     "E = POA × Pdc × ΠF_loss × Availability × (1-Curtailment) × F_other", "kWh", "전문가",
     "개별 손실 FACTOR", "PR을 동시에 적용하지 않음"],
    [16, "발전성능지수", "performance_index", "PI = Actual Energy ÷ Weather-adjusted Expected Energy",
     "%", "전체", "실제 발전량, 기상보정 예상량", "실제 발전량이 있는 월만 연간 합산"],
    [17, "실제 PR", "actual_pr", "PR = (E_actual/Pdc) ÷ (H_POA/Gref)", "%", "상세/전문가", "현장 POA",
     "현장 POA가 없으면 실제 PR로 표기하지 않음"],
    [18, "특성발전량", "specific_yield", "Yf = E / Pdc", "kWh/kWp", "전체", "발전량, DC 용량", "다른 용량 발전소 비교용"],
    [19, "설비이용률", "capacity_factor", "CF = E_actual / (Pac × 8,760)", "%", "전체", "실제 연간 발전량, AC 용량",
     "부분연도는 별도 시간 보정 필요"],
    [20, "입력기간 예상량", "expected_weather_matched_kwh", "실제 발전량이 입력된 월의 기상보정 예상량만 합산",
     "kWh", "전체", "월별 실제값 존재 여부", "부분연도를 연간 실적과 혼동하지 않음"],
    [21, "입력기간 발전성능지수", "performance_index", "ΣE_actual(입력월) ÷ ΣE_weather-adjusted(동일 입력월)",
     "%", "전체", "입력월 실제·예상", "결측월은 0으로 대체하지 않음"],
    [22, "자료 커버리지", "coverage_pct", "실제 발전량 입력 월수 ÷ 12 × 100", "%", "전체", "actual_month_count",
     "부분연도 판정 문구에 표시"],
]


def build_calc(wb):
    ws = wb.create_sheet(CALC)
    title_block(ws, "계산식 명세", "프로그램 코드(app.py)와 Excel 기준모델이 동일하게 사용해야 하는 월별 계산 순서입니다. 실제 PR은 현장 POA가 있을 때만 표시합니다.", "H")
    section(ws, 4, "■ 모델 흐름", "H")
    header_row(ws, 5, ["순서", "단계", "내용"], height=20)
    ws.merge_cells("C5:H5")
    for col in range(3, 9):
        ws.cell(row=5, column=col).fill = FILL_HEAD
        ws.cell(row=5, column=col).border = BOX
    for i, (no, step, desc) in enumerate(FLOW_ROWS):
        r = 6 + i
        for col, v in ((1, no), (2, step), (3, desc)):
            c = ws.cell(row=r, column=col, value=v)
            c.font = F_BODY
            c.alignment = AL_C if col == 1 else AL_LW
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BOX
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 18

    start = 6 + len(FLOW_ROWS) + 1
    section(ws, start, "■ 계산식 상세", "H")
    header_row(ws, start + 1, ["순서", "계산 항목", "코드 변수", "수식/알고리즘", "단위", "적용 모드", "입력/출처", "예외·주의"])
    table_body(ws, start + 2, CALC_ROWS)
    for r in range(start + 2, start + 2 + len(CALC_ROWS)):
        ws.cell(row=r, column=1).alignment = AL_C
    widths(ws, {"A": 6, "B": 20, "C": 26, "D": 56, "E": 13, "F": 12, "G": 22, "H": 30})
    ws.freeze_panes = f"B{start + 2}"
    page_setup(ws, rows_to_repeat=f"{start + 1}:{start + 1}")
    return ws


# ==========================================================================
# 05_FACTOR_기준값
# ==========================================================================
FACTOR_ROWS = [
    ["pr", "기준 PR", 0.800, "0.75~0.85", "-", "PR 기반", "정상 설비의 온도·변환·배선 등 종합손실",
     "발전소 설계 및 운영수준에 맞게 조정",
     "https://pvpmc.sandia.gov/modeling-guide/5-ac-system-output/pv-performance-metrics/performance-ratio/"],
    ["availability", "가동률", 0.995, "0.95~1.00", "-", "공통", "고장·정비로 발전 불가능한 시간",
     "PR과 별도 적용 시 정의 일관성 필요", ""],
    ["curtailment_rate", "출력제한률", 0.000, "0~0.10", "-", "공통", "계통·계약·운영 지시에 의한 의도적 감발",
     "고장·정비는 포함하지 않음", ""],
    ["other_factor", "기타보정계수", 1.000, "0.90~1.05", "-", "공통", "별도 안전계수 또는 프로젝트 특수보정",
     "PR 손실과 중복 금지", ""],
    ["annual_degradation_rate", "연간 모듈 열화율", 0.005, "모듈 보증값", "/year", "공통", "운영연수 보정",
     "실제 장기추세와 비교", ""],
    ["albedo", "지면반사율", 0.200, "0.15~0.30", "-", "POA", "지면반사 성분", "적설·밝은 지붕은 별도 검토",
     "https://pvpmc.sandia.gov/modeling-guide/1-weather-design-inputs/plane-of-array-poa-irradiance/"],
    ["iam", "입사각 FACTOR", 0.980, "0.95~1.00", "-", "항목별", "입사각 광학손실", "PR 방식에서는 중복 적용 금지", ""],
    ["shading", "음영 FACTOR", 0.990, "0.80~1.00", "-", "항목별", "주변 지형·구조물 음영", "시간대별 또는 음영해석 권장", ""],
    ["soiling", "오염 FACTOR", 0.970, "0.90~1.00", "-", "항목별", "먼지·조류분변·염해", "강우·세척 전후 성능 확인", ""],
    ["temperature", "온도 FACTOR", 0.940, "0.85~0.98", "-", "항목별", "모듈 고온 손실", "시간별 기온·풍속·모듈온도 적용 권장", ""],
    ["mismatch", "불일치 FACTOR", 0.980, "0.95~1.00", "-", "항목별", "모듈·스트링 전기적 불일치", "MPPT 데이터가 있으면 실측 진단", ""],
    ["dc_wiring", "DC 배선 FACTOR", 0.985, "0.97~1.00", "-", "항목별", "DC 케이블·접속 손실", "설계 전압강하 자료 확인", ""],
    ["inverter", "인버터 FACTOR", 0.975, "0.95~0.99", "-", "항목별", "DC-AC 변환효율", "클리핑은 별도 또는 모델 내 포함 정의 필요", ""],
    ["ac_wiring", "AC 배선 FACTOR", 0.990, "0.98~1.00", "-", "항목별", "AC 배선 손실", "계량기 위치 확인", ""],
    ["transformer", "변압기 FACTOR", 0.990, "0.97~1.00", "-", "항목별", "변압기 손실", "발전량 계측 위치에 따라 제외 가능", ""],
]


def build_factor(wb):
    ws = wb.create_sheet(FACT)
    title_block(ws, "FACTOR 기준값", "FACTOR 기본값의 단일 출처(single source) 시트입니다. 관리자 설정값으로 관리하고 분석 실행 시 스냅샷을 보관합니다.", "I")
    section(ws, 4, "■ 손실·보정 FACTOR 기본값", "I")
    header_row(ws, 5, ["코드", "한글명", "기본값", "권장범위", "단위", "적용방식", "포함/제외 범위", "근거·주의", "출처 URL"])
    table_body(ws, 6, FACTOR_ROWS)
    for i in range(len(FACTOR_ROWS)):
        r = 6 + i
        c = ws.cell(row=r, column=3)
        c.font = F_INPUT
        c.fill = FILL_INPUT
        c.alignment = AL_C
        c.number_format = NUM_3
        ws.cell(row=r, column=4).alignment = AL_C
        ws.cell(row=r, column=5).alignment = AL_C
        ws.cell(row=r, column=6).alignment = AL_C
        u = ws.cell(row=r, column=9)
        if u.value:
            u.font = Font(name=FONT, size=8.5, color="0563C1", underline="single")
            u.hyperlink = u.value

    note_r = 6 + len(FACTOR_ROWS) + 1
    for i, text in enumerate([
        "※ PR 기반 방식(간편·상세)과 항목별 손실 방식(전문가)은 동시에 적용하지 않습니다. PR에 이미 포함된 손실을 항목별로 다시 차감하면 이중 차감이 됩니다.",
        f"※ 기준 PR·가동률·지면반사율은 본 시트가 기준이며, '{REGION}' 시트의 기본값과 '{MODEL}' 시트의 기본 입력값은 이 값을 참조합니다.",
        "※ 출처가 비어 있는 항목은 프로젝트 관리자 설정값이며, 실제 설비 사양·운영 기록으로 대체하는 것을 권장합니다.",
    ]):
        c = ws.cell(row=note_r + i, column=1, value=text)
        c.font = F_NOTE
        c.alignment = AL_LW
        ws.merge_cells(start_row=note_r + i, start_column=1, end_row=note_r + i, end_column=9)
        ws.row_dimensions[note_r + i].height = 16

    widths(ws, {"A": 24, "B": 20, "C": 11, "D": 13, "E": 8, "F": 11, "G": 34, "H": 34, "I": 40})
    ws.freeze_panes = "C6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 06_진단규칙
# ==========================================================================
DIAG_ROWS = [
    ["WEATHER_LOW", "분석연도 일사량 부족", "지역/현장 GHI", "연간 GHI ÷ 10년평균 GHI < 0.96", "보통~높음",
     "분석연도와 장기평균 일사량 비율", "현장 POA, 인근 관측소 거리", "기상영향과 설비성능 차이를 분리", "구현"],
    ["PERF_LOW", "기상 보정 후 성능저하", "실제 발전량, 기상보정 예상량", "연간 PI < 0.97 (<0.92 보통, <0.85 높음)",
     "낮음~높음", "실제/예상 비율과 저하월 목록", "정지·알람·세척·출력제한", "설비 및 운영 기록 순차 점검", "구현"],
    ["INV_LOW", "특정 인버터 저하", "인버터별 DC용량, 연간발전량", "특성발전량이 중앙값 대비 -7% 이하", "보통~높음",
     "인버터별 kWh/kWp 편차", "MPPT 전류, 알람, 차단기", "인버터·스트링 현장점검", "구현"],
    ["CLIPPING", "인버터 클리핑 가능성", "DC/AC 비율", "DC/AC ≥ 1.25 (≥1.35 보통)", "낮음~보통", "DC/AC 설계비율",
     "맑은 날 시간별 AC 출력", "정오 출력 평탄화 확인", "부분 구현"],
    ["CURTAIL", "출력제한 손실", "출력제한률/이벤트", "출력제한률 > 0 (≥5% 높음)", "보통~높음", "입력된 제한률",
     "계통지시·접속용량·PPA", "이벤트별 손실량 검증", "구현"],
    ["AGE", "모듈 경년열화", "상업운전일, 열화율", "운영연수 ≥ 5년", "낮음", "운영연수와 누적 보정률",
     "기상보정 연도별 추세", "보증 열화율과 실제 추세 비교", "구현"],
    ["MISSING", "원인진단 자료 부족", "데이터 등급", "등급 C 또는 D", "보통", "부족한 데이터 목록",
     "POA/GHI, 인버터, 정지기록", "추가 계측자료 확보", "구현"],
    ["ZERO_GEN", "일사 존재 시 무발전", "시간/일별 POA와 발전량", "POA 임계값 이상 & 발전량 0", "높음", "발생시각과 지속시간",
     "계통, 인버터 상태, 통신", "고장·정전·계량 오류 구분", "후속"],
    ["MORNING_PM", "오전/오후 비대칭", "시간별 출력·POA", "정오 기준 대칭구간 편차 지속", "보통", "시간대별 정상화 출력곡선",
     "현장 음영, 방위, 센서", "음영 및 방향 불일치 확인", "후속"],
    ["SOILING", "오염 가능성", "PI, 강우/세척 이력", "강우·세척 후 PI 유의 회복", "보통", "전후 성능 변화",
     "세척일, 강우량, 사진", "오염도 점검 및 세척주기 최적화", "후속"],
    ["SENSOR_POA", "POA 센서 이상", "GHI/POA/태양기하", "POA/GHI 비율이 방향·계절과 불일치", "보통", "센서 비율·고정값·결측",
     "설치각, 교정일, 청소상태", "센서 재설치·교정", "후속"],
]


def build_diag(wb):
    ws = wb.create_sheet(DIAG)
    title_block(ws, "저하원인 진단규칙", "진단은 고장 확정이 아니라 가능성·근거·추가 확인·권장조치로 출력합니다. 필요 데이터가 없으면 '판단자료 부족'으로 반환합니다.", "I")
    section(ws, 4, "■ 진단규칙 및 임계값", "I")
    header_row(ws, 5, ["규칙코드", "원인명", "필요 데이터", "조건/임계값(MVP)", "가능성", "자동 근거", "추가 확인", "권장 조치", "구현"])
    table_body(ws, 6, DIAG_ROWS)
    for r in range(6, 6 + len(DIAG_ROWS)):
        ws.cell(row=r, column=5).alignment = AL_C
        ws.cell(row=r, column=9).alignment = AL_C
    widths(ws, {"A": 15, "B": 20, "C": 22, "D": 32, "E": 11, "F": 24, "G": 22, "H": 26, "I": 10})
    ws.freeze_panes = "C6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 07_데이터사전
# ==========================================================================
DICT_ROWS = [
    ["발전소", "plant_name", "발전소명", "string", "-", "전체", "1MW 태양광발전소", "기본명 적용", ""],
    ["발전소", "region", "지역권역", "enum", "-", "전체", "서울경기 등 10개 권역", "필수", "주소 기반 자동선정 후속"],
    ["발전소", "analysis_year", "분석연도", "integer", "year", "전체", "2016~2025", "필수", "현재 기준자료 기간"],
    ["발전소", "commercial_date", "상업운전일", "date", "YYYY-MM-DD", "선택", "2020-01-01", "열화 FACTOR=1", ""],
    ["설비", "capacity_dc_kwp", "DC 설치용량", "number", "kWp", "전체", "1000", "필수", "발전량 정규화 기준"],
    ["설비", "capacity_ac_kw", "AC 설치용량", "number", "kW", "상세", "900", "DC와 같게 추정", "클리핑/이용률 산정"],
    ["설비", "inverters", "인버터 배열", "array", "-", "B 이상", "이름, AC/DC용량, 발전량", "없음", "인버터 편차 분석"],
    ["설치조건", "tilt_deg", "모듈 경사각", "number", "degree", "C 이상", "0~90", "30° 추정", "추정값 표시"],
    ["설치조건", "azimuth_deg", "모듈 방위각", "number", "degree", "C 이상", "북0 동90 남180 서270", "180° 추정", ""],
    ["설치조건", "latitude", "현장 위도", "number", "degree N", "전문가", "33~38", "권역 대표위도", ""],
    ["발전량", "actual_monthly_kwh", "월별 실제 발전량", "array[12]", "kWh/month", "전체", "1~12월", "미입력 월 제외", "0과 결측을 구분"],
    ["일사량", "actual_ghi_monthly", "현장 월별 GHI", "array[12]", "kWh/m²/month", "B 이상", "1~12월", "선택연도 지역 GHI", ""],
    ["일사량", "actual_poa_monthly", "현장 월별 POA", "array[12]", "kWh/m²/month", "A/B", "1~12월", "GHI에서 계산", "현장값 최우선"],
    ["계산", "calculation_method", "계산방식", "enum", "-", "전체", "pr_based / component_loss_based", "pr_based",
     "항목별 방식은 전문가 권한"],
    ["계산", "component_losses", "항목별 FACTOR", "object", "-", "전문가", "iam, shading 등", "기본값", "PR과 중복 금지"],
    ["출력", "performance_index", "발전성능지수", "number", "%", "전체", "Actual / Weather expected", "실적 없으면 null", ""],
    ["출력", "expected_weather_matched_kwh", "입력기간 기상보정 예상량", "number", "kWh", "전체", "실제값이 있는 월만 합산",
     "실적 미입력 시 0", "부분연도 비교 기준"],
    ["출력", "actual_month_count", "실적 입력 월수", "integer", "month", "전체", "0~12", "0", "자료 범위 표시"],
    ["출력", "coverage_pct", "실적 자료 커버리지", "number", "%", "전체", "0~100", "0", "actual_month_count / 12"],
    ["출력", "data_quality.grade", "데이터 등급", "enum", "-", "전체", "A/B/C/D", "자동", "분석범위 제한"],
    ["출력", "diagnoses", "저하요인 진단", "array", "-", "전체", "원인, 가능성, 근거, 조치", "최소 1개", "고장 확정 금지"],
    ["전문가", "expert_evidence", "전문가 첨부자료", "array", "PNG/JPG", "전문가", "id, 파일명, 제목, 설명", "미입력 허용",
     "PPT/PDF 부록 자동 반영"],
]


def build_dict(wb):
    ws = wb.create_sheet(DICT)
    title_block(ws, "입력·출력 데이터 사전", "웹 API와 보고서에서 사용하는 주요 필드입니다. 모든 시계열 값에는 시간대·단위·출처 및 품질 플래그를 함께 관리하는 것이 목표입니다.", "I")
    section(ws, 4, "■ 필드 정의", "I")
    header_row(ws, 5, ["영역", "필드명", "한글명", "자료형", "단위", "필수 모드", "허용값/예시", "결측 처리", "비고"])
    table_body(ws, 6, DICT_ROWS)
    for r in range(6, 6 + len(DICT_ROWS)):
        for col in (1, 4, 5, 6):
            ws.cell(row=r, column=col).alignment = AL_C
    widths(ws, {"A": 11, "B": 30, "C": 22, "D": 11, "E": 14, "F": 11, "G": 28, "H": 18, "I": 24})
    ws.freeze_panes = "C6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 08_검증CASE  (구버전에서 중복되던 TC-012 를 정정하고 일련번호 재부여)
# ==========================================================================
TEST_ROWS = [
    ["정상 1MW 남향 30°", "서울경기, 2025, DC 1000, AC 1000, PR 80%, 실제=기상보정 예상", "PI 약 100%, 정상",
     "월/연 예상량, PI", "월 ±0.1%, 연 ±0.05%", "수동/코드", "기본 회귀테스트"],
    ["실제 발전량 미입력", "TC-001과 동일, 실제 12개월 공란", "'실적 입력 필요' 반환", "판정, null 처리", "오류 없음", "코드", ""],
    ["일사량이 장기평균보다 낮은 연도", "지역 연도 GHI 비율 < 96%", "기상요인 진단", "weather_ratio, 진단코드",
     "임계값 정확 일치", "코드", ""],
    ["기상보정 후 10% 저하", "실제 = 기상보정 예상 × 0.90", "관찰 또는 주의, PERF_LOW", "PI, 판정, 진단", "임계값 일치", "코드", ""],
    ["특정 인버터 10% 저하", "2대 동일 DC, 한 대 kWh/kWp -10%", "INV_LOW 보통", "중앙값 편차", "±0.1%p", "코드", ""],
    ["DC/AC 1.35", "DC 1000, AC 740", "클리핑 가능성 보통", "비율, 진단", "조건 경계 확인", "코드", "시간자료 없으므로 확정 금지"],
    ["출력제한 5%", "curtailment_rate = 0.05", "예상량 5% 감소, CURTAIL 높음", "발전량, 진단", "±0.01%", "코드", ""],
    ["상업운전 10년", "연 열화율 0.5%, 운영 10년", "F_age 약 0.951", "열화 FACTOR", "±0.001", "코드", ""],
    ["현장 POA 우선", "POA 12개월 입력, GHI도 입력", "기상예상량은 POA 직접 적용", "weather_source, 예상량", "POA 우선", "코드", ""],
    ["항목별 손실 권한", "전문가 토큰 없이 component 요청", "PR 방식으로 강제", "calculation_method", "권한 차단", "코드", ""],
    ["결측 월", "실제 6개월만 입력", "해당 6개월 예상량 합으로 PI 계산", "annual actual_month_count", "6", "코드",
     "부분연도 문구 및 동일기간 예상량 확인"],
    ["단위 오류 파일", "발전량 컬럼에 문자열/음수 포함", "유효행만 집계, 음수 제외", "rows_valid, 월합계", "오류 메시지 명확", "코드", ""],
    ["업로드 결측월", "1월·3월만 있는 CSV", "2월 등 결측월은 null, 0으로 대체하지 않음", "업로드 monthly array", "null 유지", "코드",
     "데이터품질 경고 표시"],
    ["로컬 프로젝트 저장", "입력값·분석결과 저장 후 재조회", "동일 입력·결과 복원", "SQLite CRUD", "일치", "코드", "외부 전송 없음"],
    ["전문가 이미지 부록", "전문가 토큰 + PNG/JPG 첨부", "PPT/PDF에 제목·설명과 이미지 포함", "보고서 페이지/슬라이드", "오버플로 없음",
     "코드", "전문가 권한 필요"],
]


def build_test(wb):
    ws = wb.create_sheet(TEST)
    title_block(ws, "계산 및 진단 검증 CASE", "프로그램 배포 전 동일 입력으로 Excel/코드 결과를 비교합니다. 월별 예상량 허용오차 ±0.1%, 연간 ±0.05%를 기본으로 합니다.", "H")
    section(ws, 4, "■ 검증 CASE 목록", "H")
    header_row(ws, 5, ["CASE ID", "시나리오", "핵심 입력", "예상 결과", "검증 지표", "허용오차/판정", "자동화 상태", "비고"])
    rows = [[f"TC-{i + 1:03d}"] + row for i, row in enumerate(TEST_ROWS)]
    table_body(ws, 6, rows)
    for r in range(6, 6 + len(rows)):
        ws.cell(row=r, column=1).alignment = AL_C
        ws.cell(row=r, column=1).font = F_BODY_B
        ws.cell(row=r, column=7).alignment = AL_C
    note_r = 6 + len(rows) + 1
    c = ws.cell(row=note_r, column=1,
                value="※ 구버전 자료에서 TC-012 가 두 건 중복되어 있어 v1.0에서 일련번호를 다시 부여했습니다(총 15건).")
    c.font = F_NOTE
    c.alignment = AL_LW
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=8)
    widths(ws, {"A": 11, "B": 24, "C": 36, "D": 30, "E": 24, "F": 18, "G": 12, "H": 26})
    ws.freeze_panes = "B6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 09_보고서구성
# ==========================================================================
REPORT_ROWS = [
    ["고객 PDF", "표지/분석조건", "발전소, 분석연도, 용량, 모델버전", "승인자/기관정보", "예", "구현", "전자서명·승인"],
    ["고객 PDF", "종합판정/KPI", "실제·장기·기상보정 예상, PI, 데이터등급", "전문가 요약 의견", "예", "구현", "브랜드 템플릿"],
    ["고객 PDF", "월별 비교", "월별 실제·예상 표", "특이월 주석", "예", "구현", "그래프 이미지 강화"],
    ["고객 PDF", "저하요인/권장조치", "가능성, 근거, 권장조치", "현장확인 결과", "예", "구현(전문가 이미지 부록 포함)", "사진 첨부"],
    ["고객 PPT", "표지/종합판정", "고객 발표용 핵심 결과", "설명 순서 수정", "예", "구현", "슬라이드 선택"],
    ["고객 PPT", "월별 발전량 차트", "기상보정 예상과 실제", "특이월 강조", "예", "구현", "회사 테마"],
    ["고객 PPT", "전문가 검토의견", "입력한 전문가 의견", "현장사진, 열화상, 알람, IV Curve", "예", "구현(이미지 업로드 포함)",
     "슬라이드 순서·선택·회사 테마"],
    ["상세 Excel", "분석요약/월별", "수치와 수식, 차트", "검토 의견", "전문가", "구현", "원본데이터 포함"],
    ["상세 Excel", "계산식/FACTOR", "계산변수, 손실계수, 데이터품질", "관리자 승인값", "전문가", "구현", "버전 스냅샷"],
]


def build_report(wb):
    ws = wb.create_sheet(REPORT)
    title_block(ws, "보고서 구성", "PDF는 공식 보관용, PPT는 고객 설명·전문가 편집용, Excel은 계산검증·기술분석용으로 역할을 분리합니다.", "G")
    section(ws, 4, "■ 산출물별 구성", "G")
    header_row(ws, 5, ["산출물", "섹션/슬라이드", "자동 생성내용", "전문가 추가내용", "고객 공개", "MVP 구현", "후속 기능"])
    table_body(ws, 6, REPORT_ROWS)
    for r in range(6, 6 + len(REPORT_ROWS)):
        ws.cell(row=r, column=5).alignment = AL_C
    widths(ws, {"A": 14, "B": 18, "C": 32, "D": 30, "E": 11, "F": 24, "G": 26})
    ws.freeze_panes = "C6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 10_개발연계
# ==========================================================================
DEV_ROWS = [
    ["백엔드", "app.py + solar_core.py", "API·권한·업로드 및 POA/예상량/PI/진단 계산", MODEL_VERSION,
     "pytest/API 회귀테스트", "전문가 PIN·토큰 비밀값은 배포 전 변경", "RBAC/사용자 인증"],
    ["프런트엔드", "static/index.html, app.js", "간편·상세·전문가 입력, 프로젝트 저장, 결과표시", MODEL_VERSION,
     "브라우저 입력/다운로드 테스트", "전문가 기능은 서버 토큰 검증", "Next.js 전환"],
    ["기준 데이터", "data/solar_data.json", "2016~2025 10개 권역 월별 GHI", "2016-2025",
     "원자료 1,200행 대조", "출처·기간·버전 보관", "기상 API 자동갱신"],
    ["고객 PDF", "/api/report/pdf + reports.py", "공식 고객 보고서 및 전문가 이미지 부록", MODEL_VERSION,
     "PDF 렌더링 검토", "면책·모델버전 포함", "승인 워크플로"],
    ["고객 PPT", "/api/report/pptx + PptxGenJS", "고객 설명 및 전문가 첨부자료 슬라이드", MODEL_VERSION,
     "슬라이드 렌더·오버플로 검사", "자동결과와 전문가 의견 구분", "브랜드 템플릿"],
    ["상세 Excel", "/api/report/excel + reports.py", "월별 결과, FACTOR, 계산식, 데이터품질", MODEL_VERSION,
     "LibreOffice 재계산·수식 오류 검사", "동적배열 함수 사용 금지", "원본자료 상세"],
    ["전문가 권한", "/api/expert/unlock", "만료시간이 있는 서명 토큰 발급", MODEL_VERSION,
     "잘못된 PIN 401, 만료/변조 토큰 확인", "초기 PIN은 데모 전용이므로 배포 전 반드시 변경", "RBAC/OAuth"],
    ["모델 변경", "MODEL_VERSION", "보고서·결과 JSON·PPT/Excel에 기록", MODEL_VERSION,
     "코드·Excel 버전 일치 확인", "분석 실행 시 입력·결과 스냅샷", "모델 레지스트리"],
    ["로컬 저장", "storage.py / SQLite", "프로젝트 입력값·분석결과 저장·불러오기", MODEL_VERSION,
     "CRUD API 테스트", "PC 로컬 파일 권한·백업 필요", "다중 사용자 DB"],
    ["전문가 첨부", "/api/evidence/*", "현장사진·열화상·알람 이미지 로컬 저장", MODEL_VERSION,
     "권한·용량·확장자 테스트", "PNG/JPG, 10MB 제한", "문서/IV Curve 파일 지원"],
]


def build_dev(wb):
    ws = wb.create_sheet(DEV)
    title_block(ws, "웹 프로그램 개발 연계", "현재 MVP 코드와 본 워크북을 같은 입력으로 비교하고, 변경 시 모델 버전과 변경이력을 함께 갱신합니다.", "G")
    section(ws, 4, "■ 컴포넌트별 연계", "G")
    header_row(ws, 5, ["구분", "경로/컴포넌트", "역할", "현재 버전", "검증방법", "보안/운영 주의", "후속"])
    table_body(ws, 6, DEV_ROWS)
    for r in range(6, 6 + len(DEV_ROWS)):
        ws.cell(row=r, column=4).alignment = AL_C
    widths(ws, {"A": 13, "B": 28, "C": 34, "D": 13, "E": 26, "F": 30, "G": 20})
    ws.freeze_panes = "C6"
    page_setup(ws, rows_to_repeat="5:5")
    return ws


# ==========================================================================
# 11_발전량_모델  (셀 주소는 다른 시트가 참조하므로 구버전과 동일하게 유지)
# ==========================================================================
MODEL_INPUTS = [
    (5, "분석 지역", "서울경기", None, "광역권 10개 중 선택"),
    (6, "분석 기준", "10년평균", None, "10년평균 또는 개별연도(2016~2025)"),
    (7, "경사각 β (°)", 30, "0.0", "0° = 수평, 90° = 수직"),
    (8, "방위각 A (°)", 180, "0.0", "0° 북 / 90° 동 / 180° 남 / 270° 서"),
    (9, "지면반사율 ρg", 0.2, NUM_3, f"일반 토지 기본 0.20 (기본값 출처: {FACT})"),
    (10, "DC 설비용량 (kWp)", 1000, NUM_INT, "발전량·특성발전량 환산 기준"),
    (11, "기준 PR", 0.8, NUM_3, "온도·인버터·배선·오염 등 종합손실 포함"),
    (12, "가동률", 0.995, NUM_3, "고장·정비로 발전 불가능한 시간 반영"),
    (13, "출력제한/계통손실률", 0, NUM_3, "의도적 감발(고장·정비 제외)"),
    (14, "기타 보정계수", 1, NUM_3, "프로젝트 특수보정 (PR과 중복 금지)"),
]

MODEL_RESULTS = [
    (5, "연간 수평면 일사량", "=SUM(C22:C33)", "kWh/m²/년", NUM_1),
    (6, "연간 경사면 일사량 POA", "=SUM(P22:P33)", "kWh/m²/년", NUM_1),
    (7, "경사면/수평면 FACTOR", "=IF(F5=0,0,F6/F5)", "배", NUM_3),
    (8, "연간 특성발전량", "=SUM(R22:R33)", "kWh/kWp", NUM_1),
    (9, "연간 발전량", "=SUM(S22:S33)", "MWh/년", NUM_1),
    (10, "설비이용률", "=IF(B10=0,0,F9/(B10*8.76))", "%", NUM_PCT),
    (11, "경사면 증가율", "=IF(F5=0,0,F6/F5-1)", "%", NUM_PCT),
    (12, "검토 상태", '=IF(COUNTIF(T22:T33,"<>정상")=0,"정상","검토 필요")', "-", None),
]

MODEL_NOTES = [
    "① 월 수평면 일사량(원자료)을 월 일평균 H로 환산합니다.",
    "② 대기권외 일사량 H0와 청명도지수 Kt = H/H0 를 계산합니다.",
    "③ Erbs 월평균 상관식으로 산란 성분 Hd 를 추정합니다.",
    "④ 15분 태양기하 적분으로 방위·경사 직달계수 Rb 를 계산합니다.",
    "⑤ POA = Hb×Rb + Hd×(1+cosβ)/2 + H×ρg×(1-cosβ)/2",
    "⑥ 발전량 = POA × DC용량 × PR × 가동률 × (1-출력제한) × 기타계수",
    "",
    "※ 광역권 월자료 기반의 예비·타당성 분석용입니다.",
    "※ 금융·성능보증용은 현장 좌표, 시간별 DNI/DHI, 음영, 온도, 출력제한 자료를 적용하세요.",
    f"※ 노란 배경·파란 글씨 셀만 입력값이며, 상세 한계는 '{LIMIT}' 시트를 참고하세요.",
]

MODEL_TABLE_HEAD = [
    "월", "일수", "수평면\n월일사량", "수평면\n일평균 H", "대기권외\nH0", "Kt", "Kt\n적용값", "산란비\nHd/H",
    "산란 H_d", "직달 H_b", "Rb", "직달 POA", "천공산란\nPOA", "지면반사\nPOA", "POA\n일평균", "POA\n월일사량",
    "POA\nFACTOR", "특성발전량", "발전량", "검토",
]
MODEL_TABLE_UNIT = [
    "-", "일", "kWh/m²", "kWh/m²/일", "kWh/m²/일", "-", "-", "-", "kWh/m²/일", "kWh/m²/일", "-",
    "kWh/m²/일", "kWh/m²/일", "kWh/m²/일", "kWh/m²/일", "kWh/m²", "배", "kWh/kWp", "MWh", "-",
]
MODEL_TABLE_FMT = [
    "0", NUM_1, NUM_1, NUM_2, NUM_2, NUM_3, NUM_3, NUM_3, NUM_2, NUM_2, NUM_3, NUM_3, NUM_3, NUM_3,
    NUM_2, NUM_1, NUM_3, NUM_1, NUM_1, None,
]


def build_model(wb):
    ws = wb.create_sheet(MODEL)
    title_block(ws, "경사면 일사량 및 발전량 계산 모델",
                "지역별 월 수평면 일사량을 직달·산란으로 분리하고, 경사각·방위각·지면반사를 반영한 POA와 발전량을 계산합니다.", "T")

    def band(row, c1, c2, text):
        ws.cell(row=row, column=c1, value=text).font = F_SECTION
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.row_dimensions[row].height = 20

    band(4, 1, 4, "① 분석 입력")
    band(4, 5, 7, "② 주요 결과")
    band(4, 9, 14, "③ 계산 절차 및 주의사항")

    for row, label, value, fmt, desc in MODEL_INPUTS:
        c = ws.cell(row=row, column=1, value=label)
        c.font = F_BODY_B
        c.alignment = AL_L
        c.border = BOX
        input_cell(ws, f"B{row}", value, fmt)
        d = ws.cell(row=row, column=3, value=desc)
        d.font = F_NOTE
        d.alignment = AL_LW
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        for col in (3, 4):
            ws.cell(row=row, column=col).border = BOX
        ws.row_dimensions[row].height = 18

    lat_rows = [
        (15, "지역 대표 위도 (°N)", f"=INDEX({q(REGION)}$C$5:$C$14,MATCH(B5,{q(REGION)}$A$5:$A$14,0))",
         "선택 지역의 대표 도시 위도(자동)", "link"),
        (16, "현장 위도 수동입력 (선택)", None, "실제 발전소 위도를 알면 입력(권장)", "input"),
        (17, "적용 위도 φ (°N)", '=IF(B16="",B15,B16)', "수동입력이 있으면 우선 적용", "calc"),
    ]
    for row, label, formula, desc, kind in lat_rows:
        c = ws.cell(row=row, column=1, value=label)
        c.font = F_BODY_B
        c.alignment = AL_L
        c.border = BOX
        v = ws.cell(row=row, column=2, value=formula)
        v.alignment = AL_C
        v.border = BOX
        v.number_format = "0.0000"
        if kind == "input":
            v.font = F_INPUT
            v.fill = FILL_INPUT
        elif kind == "link":
            v.font = F_LINK
        else:
            v.font = F_BODY_B
        d = ws.cell(row=row, column=3, value=desc)
        d.font = F_NOTE
        d.alignment = AL_LW
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        for col in (3, 4):
            ws.cell(row=row, column=col).border = BOX
        ws.row_dimensions[row].height = 18

    for row, label, formula, unit, fmt in MODEL_RESULTS:
        c = ws.cell(row=row, column=5, value=label)
        c.font = F_BODY_B
        c.alignment = AL_L
        c.border = BOX
        v = ws.cell(row=row, column=6, value=formula)
        v.font = F_KPI
        v.fill = FILL_KPI
        v.alignment = AL_R
        v.border = BOX
        if fmt:
            v.number_format = fmt
        else:
            v.alignment = AL_C
        u = ws.cell(row=row, column=7, value=unit)
        u.font = F_NOTE
        u.alignment = AL_C
        u.border = BOX

    for i, text in enumerate(MODEL_NOTES):
        r = 5 + i
        c = ws.cell(row=r, column=9, value=text)
        c.font = F_NOTE
        c.alignment = AL_LW
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=14)

    band(20, 1, 20, "④ 월별 계산 결과 (1~12월)")
    for i, (h, u) in enumerate(zip(MODEL_TABLE_HEAD, MODEL_TABLE_UNIT)):
        c = ws.cell(row=21, column=1 + i, value=h if u == "-" else f"{h}\n[{u}]")
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_CW
        c.border = BOX
    ws.row_dimensions[21].height = 42

    geo, reg = q(GEO), q(REGION)
    for m in range(1, 13):
        r = 21 + m
        f = {
            1: m,
            2: f'=IF($B$6="10년평균",AVERAGEIFS(RawDataKWh[일수],RawDataKWh[월],$A{r},RawDataKWh[지역명],$B$5),'
               f'SUMIFS(RawDataKWh[일수],RawDataKWh[월],$A{r},RawDataKWh[연도],VALUE($B$6),RawDataKWh[지역명],$B$5))',
            3: f'=IF($B$6="10년평균",AVERAGEIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[월],$A{r},RawDataKWh[지역명],$B$5),'
               f'SUMIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[월],$A{r},RawDataKWh[연도],VALUE($B$6),RawDataKWh[지역명],$B$5))',
            4: f"=IF(B{r}=0,0,C{r}/B{r})",
            5: f"=INDEX({geo}$H$4:$H$15,MATCH(A{r},{geo}$A$4:$A$15,0))",
            6: f"=IF(E{r}=0,0,D{r}/E{r})",
            7: f"=MAX({reg}$H$7,MIN({reg}$H$8,F{r}))",
            8: f"=MAX(0.05,MIN(1,IF(INDEX({geo}$D$4:$D$15,MATCH(A{r},{geo}$A$4:$A$15,0))<=81.4,"
               f"1.391-3.56*G{r}+4.189*G{r}^2-2.137*G{r}^3,1.311-3.022*G{r}+3.427*G{r}^2-1.821*G{r}^3)))",
            9: f"=D{r}*H{r}",
            10: f"=D{r}-I{r}",
            11: f"=INDEX({geo}$G$4:$G$15,MATCH(A{r},{geo}$A$4:$A$15,0))",
            12: f"=J{r}*K{r}",
            13: f"=I{r}*(1+COS(RADIANS($B$7)))/2",
            14: f"=D{r}*$B$9*(1-COS(RADIANS($B$7)))/2",
            15: f"=L{r}+M{r}+N{r}",
            16: f"=O{r}*B{r}",
            17: f"=IF(C{r}=0,0,P{r}/C{r})",
            18: f"=P{r}*$B$11*$B$12*(1-$B$13)*$B$14",
            19: f"=R{r}*$B$10/1000",
            20: f'=IF(OR(F{r}<{reg}$H$7,F{r}>{reg}$H$8),"Kt 범위 검토",IF(OR(K{r}<0,K{r}>5),"Rb 검토","정상"))',
        }
        for col in range(1, 21):
            c = ws.cell(row=r, column=col, value=f[col])
            c.font = F_BODY
            c.border = BOX
            c.alignment = AL_C
            if MODEL_TABLE_FMT[col - 1]:
                c.number_format = MODEL_TABLE_FMT[col - 1]
            if m % 2 == 0:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 17

    totals = {
        1: "연간/평균", 2: "=SUM(B22:B33)", 3: "=SUM(C22:C33)", 4: "=C34/B34", 5: "=AVERAGE(E22:E33)",
        6: "=AVERAGE(F22:F33)", 7: "=AVERAGE(G22:G33)", 8: "=SUMPRODUCT(C22:C33,H22:H33)/SUM(C22:C33)",
        9: "=SUMPRODUCT(B22:B33,I22:I33)/SUM(B22:B33)", 10: "=SUMPRODUCT(B22:B33,J22:J33)/SUM(B22:B33)",
        11: "=SUMPRODUCT(C22:C33,K22:K33)/SUM(C22:C33)", 12: "=SUMPRODUCT(B22:B33,L22:L33)/SUM(B22:B33)",
        13: "=SUMPRODUCT(B22:B33,M22:M33)/SUM(B22:B33)", 14: "=SUMPRODUCT(B22:B33,N22:N33)/SUM(B22:B33)",
        15: "=SUM(P22:P33)/SUM(B22:B33)", 16: "=SUM(P22:P33)", 17: "=P34/C34", 18: "=SUM(R22:R33)",
        19: "=SUM(S22:S33)", 20: '=IF(COUNTIF(T22:T33,"<>정상")=0,"정상","검토 필요")',
    }
    for col in range(1, 21):
        c = ws.cell(row=34, column=col, value=totals[col])
        c.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = AL_C
        c.border = BOX
        if MODEL_TABLE_FMT[col - 1]:
            c.number_format = MODEL_TABLE_FMT[col - 1]
    ws.row_dimensions[34].height = 20

    ws.conditional_formatting.add(
        "T22:T34",
        CellIsRule(operator="notEqual", formula=['"정상"'], font=Font(name=FONT, size=9.5, bold=True, color="C00000")),
    )

    dv_region = DataValidation(type="list", formula1=f"={q(REGION)}$A$5:$A$14", allow_blank=False)
    dv_basis = DataValidation(type="list", formula1=f"={q(REGION)}$M$5:$M$15", allow_blank=False)
    dv_tilt = DataValidation(type="decimal", operator="between", formula1="0", formula2="90",
                             error="경사각은 0~90° 범위로 입력하세요.", errorTitle="입력 범위 오류")
    dv_azi = DataValidation(type="decimal", operator="between", formula1="0", formula2="360",
                            error="방위각은 0~360° 범위로 입력하세요.", errorTitle="입력 범위 오류")
    dv_ratio = DataValidation(type="decimal", operator="between", formula1="0", formula2="1.05",
                              error="0~1.05 범위의 계수로 입력하세요.", errorTitle="입력 범위 오류")
    for dv in (dv_region, dv_basis, dv_tilt, dv_azi, dv_ratio):
        ws.add_data_validation(dv)
    dv_region.add(ws["B5"])
    dv_basis.add(ws["B6"])
    dv_tilt.add(ws["B7"])
    dv_azi.add(ws["B8"])
    for addr in ("B9", "B11", "B12", "B13", "B14"):
        dv_ratio.add(ws[addr])

    widths(ws, {"A": 21, "B": 14, "C": 13, "D": 13, "E": 22, "F": 13, "G": 11, "H": 12, "I": 12, "J": 12,
                "K": 11, "L": 12, "M": 12, "N": 12, "O": 12, "P": 13, "Q": 11, "R": 13, "S": 12, "T": 13})
    ws.freeze_panes = "A22"
    page_setup(ws)
    ws.sheet_properties.tabColor = BLUE
    return ws


# ==========================================================================
# 12_월별FACTOR
# ==========================================================================
MFACT_INPUTS = [
    (5, "지역", f"={q(MODEL)}B5", "-", "선택 수평면 일사량 권역"),
    (6, "분석 기준", f"={q(MODEL)}B6", "-", "10년평균 또는 개별연도"),
    (7, "위도 φ", f"={q(MODEL)}B17", "°N", "태양위치 및 H0 계산"),
    (8, "경사각 β", f"={q(MODEL)}B7", "°", "모듈면 경사"),
    (9, "방위각 A", f"={q(MODEL)}B8", "°", "북0 / 동90 / 남180 / 서270"),
    (10, "지면반사율 ρg", f"={q(MODEL)}B9", "-", "지면반사 성분"),
    (11, "기준 PR", f"={q(MODEL)}B11", "-", "종합 시스템 성능계수"),
    (12, "가동률", f"={q(MODEL)}B12", "-", "가동 가능률"),
    (13, "출력제한률", f"={q(MODEL)}B13", "-", "계통/출력 제한"),
    (14, "기타 보정", f"={q(MODEL)}B14", "-", "추가 보정계수"),
]

MFACT_HEAD = ["월", "대표일\n(DOY)", "H 수평면\n일평균", "H0 대기권외", "Kt", "Kt 적용값", "산란비", "직달비", "Rb",
              "천공 View\nFactor", "지면 View\nFactor", "POA\nFACTOR", "수평면\n월일사량", "경사면\n월일사량",
              "특성발전량", "검토"]
MFACT_FMT = ["0", "0", NUM_2, NUM_2, NUM_3, NUM_3, NUM_3, NUM_3, NUM_3, NUM_3, NUM_3, NUM_3, NUM_1, NUM_1, NUM_1, None]

MFACT_DEF = [
    ["H", "수평면 월일사량 ÷ 일수", "kWh/m²/일", "지역별 원자료"],
    ["H0", "(24×60/π)·Gsc·E0·[cosφ cosδ sinωs + ωs sinφ sinδ]", "kWh/m²/일", "대기권외 일사량"],
    ["Kt", "H / H0", "-", "대기 투과 수준"],
    ["Hd/H", "Erbs 월평균 상관식", "-", "산란 성분 비율"],
    ["Rb", "Σmax(cosAOI,0) / Σmax(cosθz,0)", "-", "직달 성분의 경사·방위 기하계수"],
    ["F_sky", "(1+cosβ)/2", "-", "등방성 천공산란 View Factor"],
    ["F_ground", "(1-cosβ)/2", "-", "지면 View Factor"],
    ["POA", "Hb×Rb + Hd×F_sky + H×ρg×F_ground", "kWh/m²/일", "경사면 일사량"],
    ["POA FACTOR", "POA 월일사량 / 수평면 월일사량", "배", "경사·방위 효과"],
    ["특성발전량", "POA × PR × 가동률 × (1-출력제한) × 기타", "kWh/kWp", "DC 1 kWp 기준"],
]


def build_mfact(wb):
    ws = wb.create_sheet(MFACT)
    title_block(ws, "월별 경사면 일사량 계산 FACTOR",
                f"'{MODEL}' 시트에서 사용되는 주요 계수와 월별 값을 한 화면에서 확인합니다. 모든 값은 참조식이며 직접 입력하지 않습니다.", "P")
    section(ws, 4, "■ 현재 적용 입력 FACTOR", "P")
    for row, label, formula, unit, desc in MFACT_INPUTS:
        c = ws.cell(row=row, column=1, value=label)
        c.font = F_BODY_B
        c.alignment = AL_L
        c.border = BOX
        v = ws.cell(row=row, column=2, value=formula)
        v.font = F_LINK
        v.alignment = AL_C
        v.border = BOX
        v.number_format = NUM_3 if unit == "-" else "0.00"
        u = ws.cell(row=row, column=3, value=unit)
        u.font = F_NOTE
        u.alignment = AL_C
        u.border = BOX
        d = ws.cell(row=row, column=4, value=desc)
        d.font = F_NOTE
        d.alignment = AL_LW
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        for col in range(4, 7):
            ws.cell(row=row, column=col).border = BOX
        ws.row_dimensions[row].height = 17

    section(ws, 17, "■ 월별 계산계수", "P")
    for i, h in enumerate(MFACT_HEAD):
        c = ws.cell(row=18, column=1 + i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_CW
        c.border = BOX
    ws.row_dimensions[18].height = 34

    geo, model = q(GEO), q(MODEL)
    for m in range(1, 13):
        r = 18 + m
        src = 21 + m  # 발전량_모델의 해당 월 행
        vals = {
            1: m,
            2: f"=INDEX({geo}$B$4:$B$15,MATCH(A{r},{geo}$A$4:$A$15,0))",
            3: f"={model}D{src}",
            4: f"={model}E{src}",
            5: f"={model}F{src}",
            6: f"={model}G{src}",
            7: f"={model}H{src}",
            8: f"=1-G{r}",
            9: f"={model}K{src}",
            10: f"=(1+COS(RADIANS({model}$B$7)))/2",
            11: f"=(1-COS(RADIANS({model}$B$7)))/2",
            12: f"={model}Q{src}",
            13: f"={model}C{src}",
            14: f"={model}P{src}",
            15: f"={model}R{src}",
            16: f"={model}T{src}",
        }
        for col in range(1, 17):
            c = ws.cell(row=r, column=col, value=vals[col])
            c.font = F_BODY if col in (1, 2, 8, 10, 11) else F_LINK
            c.alignment = AL_C
            c.border = BOX
            if MFACT_FMT[col - 1]:
                c.number_format = MFACT_FMT[col - 1]
            if m % 2 == 0:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 17

    section(ws, 33, "■ FACTOR 정의 및 계산식", "P")
    header_row(ws, 34, ["기호", "계산식", "단위", "설명"], height=20)
    ws.merge_cells("D34:P34")
    for col in range(4, 17):
        ws.cell(row=34, column=col).fill = FILL_HEAD
        ws.cell(row=34, column=col).border = BOX
    for i, row in enumerate(MFACT_DEF):
        r = 35 + i
        for col, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = F_BODY
            c.alignment = AL_C if col == 3 else AL_L
            c.border = BOX
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=16)
        for col in range(4, 17):
            ws.cell(row=r, column=col).border = BOX
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        if i % 2 == 1:
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 17

    widths(ws, {"A": 14, "B": 13, "C": 13, "D": 13, "E": 11, "F": 11, "G": 11, "H": 11, "I": 11, "J": 12,
                "K": 12, "L": 11, "M": 13, "N": 13, "O": 13, "P": 13})
    ws.freeze_panes = "B19"
    page_setup(ws)
    ws.sheet_properties.tabColor = BLUE
    return ws


# ==========================================================================
# 13_시나리오비교 + _시나리오계산(숨김)
# ==========================================================================
SCENARIOS = [
    ("현재 입력", None, None, f"'{MODEL}' 시트 입력값"),
    ("남향 15°", 15, 180, "저경사 남향"),
    ("남향 20°", 20, 180, "저경사 남향"),
    ("남향 25°", 25, 180, "국내 일반 고정식 예시"),
    ("남향 30°", 30, 180, "중경사 남향"),
    ("남향 35°", 35, 180, "고경사 남향"),
    ("남동향 25°", 25, 135, "오전 발전 비중 확대"),
    ("남서향 25°", 25, 225, "오후 발전 비중 확대"),
    ("동향 20°", 20, 90, "동향 예시"),
    ("서향 20°", 20, 270, "서향 예시"),
]

HELP_HEAD = ["시나리오ID", "시나리오", "월", "일수", "GHI 월", "GHI 일평균", "H0", "Kt", "Kt 적용", "산란비",
             "Hd", "Hb", "Rb", "POA 일평균", "POA 월", "특성발전량", "발전량(MWh)"]


def build_scenario(wb):
    ws = wb.create_sheet(SCEN)
    title_block(ws, "경사각·방위각 시나리오 비교",
                f"지역·분석기준·PR 등은 '{MODEL}' 입력을 공통으로 사용하며, 시나리오별 경사각과 방위각만 바꾸어 비교합니다.", "J")
    section(ws, 3, "■ 시나리오별 연간 결과", "J")
    header_row(ws, 4, ["시나리오", "경사각 β\n[°]", "방위각 A\n[°]", "연간 수평면\n[kWh/m²]", "연간 POA\n[kWh/m²]",
                       "POA FACTOR\n[배]", "특성발전량\n[kWh/kWp]", "발전량\n[MWh]", "기본 대비\n[%]", "설명"], height=36)

    helper = q(HELP)
    for i, (name, tilt, azi, desc) in enumerate(SCENARIOS):
        r = 5 + i
        sid = i + 1
        vals = {
            1: name,
            2: f"={q(MODEL)}B7" if tilt is None else tilt,
            3: f"={q(MODEL)}B8" if azi is None else azi,
            4: f"={q(MODEL)}F5",
            5: f"=SUMIFS({helper}$O$2:$O$121,{helper}$A$2:$A$121,{sid})",
            6: f"=IF(D{r}=0,0,E{r}/D{r})",
            7: f"=SUMIFS({helper}$P$2:$P$121,{helper}$A$2:$A$121,{sid})",
            8: f"=SUMIFS({helper}$Q$2:$Q$121,{helper}$A$2:$A$121,{sid})",
            9: f"=IF($H$5=0,0,H{r}/$H$5-1)",
            10: desc,
        }
        fmts = {2: "0", 3: "0", 4: NUM_1, 5: NUM_1, 6: NUM_3, 7: NUM_1, 8: NUM_1, 9: NUM_PCT}
        for col in range(1, 11):
            c = ws.cell(row=r, column=col, value=vals[col])
            c.font = F_BODY_B if (col == 1 or i == 0) else F_BODY
            c.alignment = AL_L if col in (1, 10) else AL_C
            c.border = BOX
            if col in fmts:
                c.number_format = fmts[col]
            if i == 0:
                c.fill = FILL_KPI
            elif i % 2 == 1:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 18

    ws.conditional_formatting.add(
        "I6:I14",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(name=FONT, size=9.5, color="C00000")),
    )
    note = ws.cell(row=16, column=1,
                   value="※ '기본 대비'는 첫 행(현재 입력) 발전량 대비 증감률입니다. 지역·PR·용량 등 나머지 조건은 모든 시나리오가 동일합니다.")
    note.font = F_NOTE
    note.alignment = AL_LW
    ws.merge_cells("A16:J16")

    chart = BarChart()
    chart.type = "col"
    chart.title = "시나리오별 연간 특성발전량 (kWh/kWp)"
    chart.height, chart.width = 8.5, 22
    chart.y_axis.title = "kWh/kWp"
    data = Reference(ws, min_col=7, min_row=4, max_row=14)
    cats = Reference(ws, min_col=1, min_row=5, max_row=14)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "A18")

    widths(ws, {"A": 16, "B": 11, "C": 11, "D": 14, "E": 14, "F": 12, "G": 14, "H": 12, "I": 12, "J": 26})
    ws.freeze_panes = "A5"
    page_setup(ws)
    ws.sheet_properties.tabColor = BLUE
    return ws


def build_scenario_helper(wb):
    ws = wb.create_sheet(HELP)
    ws.sheet_state = "hidden"
    header_row(ws, 1, HELP_HEAD, height=28)
    geo, model, scen = q(GEO), q(MODEL), q(SCEN)
    r = 2
    for i, (_name, tilt, azi, _desc) in enumerate(SCENARIOS):
        sid = i + 1
        s_row = 5 + i          # 13_시나리오비교의 시나리오 행
        b = f"{scen}$B${s_row}"  # 경사각 (POA 계산의 View Factor 에 사용)
        # 시나리오별 cosAOI⁺ 는 15_태양기하 시트에서 1회 계산되므로 여기서는 SUMIFS 로 집계만 한다.
        aoi_col = get_column_letter(16 + i)
        for m in range(1, 13):
            src = 21 + m
            rb = (f"=IF(SUMIFS({geo}$N$21:$N$1172,{geo}$A$21:$A$1172,C{r})=0,0,"
                  f"SUMIFS({geo}${aoi_col}$21:${aoi_col}$1172,{geo}$A$21:$A$1172,C{r})"
                  f"/SUMIFS({geo}$N$21:$N$1172,{geo}$A$21:$A$1172,C{r}))")
            vals = {
                1: sid,
                2: f"={scen}$A${s_row}",
                3: m,
                4: f"={model}B{src}",
                5: f"={model}C{src}",
                6: f"={model}D{src}",
                7: f"={model}E{src}",
                8: f"={model}F{src}",
                9: f"={model}G{src}",
                10: f"={model}H{src}",
                11: f"={model}I{src}",
                12: f"={model}J{src}",
                13: rb,
                14: f"=L{r}*M{r}+K{r}*(1+COS(RADIANS({b})))/2+F{r}*{model}$B$9*(1-COS(RADIANS({b})))/2",
                15: f"=N{r}*D{r}",
                16: f"=O{r}*{model}$B$11*{model}$B$12*(1-{model}$B$13)*{model}$B$14",
                17: f"=P{r}*{model}$B$10/1000",
            }
            for col in range(1, 18):
                c = ws.cell(row=r, column=col, value=vals[col])
                c.font = F_BODY
                c.alignment = AL_C
            r += 1
    ws.cell(row=r + 1, column=1,
            value=f"※ '{SCEN}' 시트의 SUMIFS 집계를 위한 보조 계산 시트입니다. 직접 편집하지 마세요.").font = F_NOTE
    widths(ws, {get_column_letter(i): 13 for i in range(1, 18)})
    ws.freeze_panes = "D2"
    return ws


# ==========================================================================
# 14_지역·상수
# ==========================================================================
REGION_ROWS = [
    ["서울경기", "서울", 37.5665, "서울 좌표를 대표값으로 적용", "현장 위도 입력 권장"],
    ["강원영동", "강릉", 37.7519, "영동권 대표 도시", "해안/산악 지형 차이 유의"],
    ["강원영서", "춘천", 37.8813, "영서권 대표 도시", "현장 위도 입력 권장"],
    ["충북", "청주", 36.6424, "충북 대표 도시", "현장 위도 입력 권장"],
    ["충남", "대전", 36.3504, "충남권 대표값", "대전은 행정구역상 별도이나 위도 대표값으로 사용"],
    ["경북", "대구", 35.8714, "경북권 대표값", "대구는 행정구역상 별도이나 위도 대표값으로 사용"],
    ["경남", "부산", 35.1796, "경남권 대표값", "부산은 행정구역상 별도이나 위도 대표값으로 사용"],
    ["전북", "전주", 35.8242, "전북 대표 도시", "현장 위도 입력 권장"],
    ["전남", "광주", 35.1595, "전남권 대표값", "광주는 행정구역상 별도이나 위도 대표값으로 사용"],
    ["제주", "제주", 33.4996, "제주권 대표 도시", "제주 남북부 차이 유의"],
]

CONST_ROWS = [
    ("태양상수 Gsc", 0.082, "MJ/m²/min", "일별 대기권외 수평면 일사량 계산", "고정", None),
    ("시간 간격", 0.25, "hour", "Rb 적분용 15분 간격", "고정", None),
    ("Erbs Kt 하한", 0.30, "-", "월평균 확산분리식 적용 범위 보정", "고정", None),
    ("Erbs Kt 상한", 0.80, "-", "월평균 확산분리식 적용 범위 보정", "고정", None),
    ("기본 지면반사율", None, "-", "일반 토지/잔디 가정", f"{FACT} 참조", f"={q(FACT)}C11"),
    ("기본 PR", None, "-", "온도·인버터·배선·오염·불일치 등 통합", f"{FACT} 참조", f"={q(FACT)}C6"),
    ("기본 가동률", None, "-", "설비 및 계통 가동 가능률", f"{FACT} 참조", f"={q(FACT)}C7"),
    ("기준 일사강도", 1, "kW/m²", "특성발전량 환산 기준", "고정", None),
]


def build_region(wb):
    ws = wb.create_sheet(REGION)
    title_block(ws, "지역 대표 위도 및 모델 상수",
                "지역별 수평면 일사량은 광역권 평균이므로 대표 위도는 초기 검토용입니다. 실제 발전소 위도로 보정하는 것을 권장합니다.", "M")
    section(ws, 3, "■ 지역 대표 위도 / ■ 모델 상수·기본값 / ■ 분석 기준 목록", "M")
    header_row(ws, 4, ["지역명", "대표 도시", "대표 위도\n[°N]", "위도 적용 설명", "비고"], height=32)
    header_row(ws, 4, ["상수/기본값", "값", "단위", "설명", "수정 권장"], start_col=7, height=32)
    hd = ws.cell(row=4, column=13, value="분석 기준 목록")
    hd.font = F_HEAD
    hd.fill = FILL_HEAD
    hd.alignment = AL_CW
    hd.border = BOX

    table_body(ws, 5, REGION_ROWS, band=True)
    for i in range(len(REGION_ROWS)):
        r = 5 + i
        ws.cell(row=r, column=2).alignment = AL_C
        c = ws.cell(row=r, column=3)
        c.alignment = AL_C
        c.number_format = "0.0000"
        ws.row_dimensions[r].height = 17

    for i, (label, value, unit, desc, memo, formula) in enumerate(CONST_ROWS):
        r = 5 + i
        cells = [(7, label), (8, formula if formula else value), (9, unit), (10, desc), (11, memo)]
        for col, v in cells:
            c = ws.cell(row=r, column=col, value=v)
            c.border = BOX
            c.alignment = AL_L if col in (7, 10, 11) else AL_C
            if col == 8:
                c.number_format = NUM_3
                c.font = F_LINK if formula else F_BODY_B
            else:
                c.font = F_BODY if col != 7 else F_BODY_B
            if i % 2 == 1:
                c.fill = FILL_BAND

    basis = ["10년평균"] + [str(y) for y in YEARS]
    for i, v in enumerate(basis):
        c = ws.cell(row=5 + i, column=13, value=int(v) if v.isdigit() else v)
        c.font = F_BODY
        c.alignment = AL_C
        c.border = BOX
        c.number_format = "0"
        if i % 2 == 1:
            c.fill = FILL_BAND

    note_r = 17
    for i, text in enumerate([
        f"※ 지면반사율·PR·가동률의 기본값은 '{FACT}' 시트를 참조합니다(값을 두 곳에서 따로 관리하지 않습니다).",
        f"※ 지역명 목록(A5:A14)과 분석 기준 목록(M5:M15)은 '{MODEL}' 시트 입력 셀의 선택 목록으로 사용됩니다.",
    ]):
        c = ws.cell(row=note_r + i, column=1, value=text)
        c.font = F_NOTE
        c.alignment = AL_LW
        ws.merge_cells(start_row=note_r + i, start_column=1, end_row=note_r + i, end_column=13)

    widths(ws, {"A": 13, "B": 11, "C": 12, "D": 30, "E": 40, "F": 2, "G": 18, "H": 11, "I": 12, "J": 34,
                "K": 18, "L": 2, "M": 15})
    ws.freeze_panes = "A5"
    page_setup(ws)
    ws.sheet_properties.tabColor = BLUE
    return ws


# ==========================================================================
# 15_태양기하
# ==========================================================================
def build_geo(wb):
    ws = wb.create_sheet(GEO)
    title_block(ws, "태양기하 및 방위·경사 FACTOR 산출",
                "월 대표일 기준 15분 간격 태양벡터를 적분하여 월별 Rb(경사면/수평면 직달 기하계수)와 대기권외 일사량 H0를 계산합니다.", "O")
    header_row(ws, 3, ["월", "대표일\n(DOY)", "적위 δ\n[°]", "일몰시간각 ωs\n[°]", "Σ cosθz", "Σ cosAOI", "Rb",
                       "H0\n[kWh/m²/일]"], height=32)
    model, reg = q(MODEL), q(REGION)
    for m in range(1, 13):
        r = 3 + m
        vals = {
            1: m,
            2: MONTH_DOY[m - 1],
            3: f"=23.45*SIN(RADIANS(360*(284+B{r})/365))",
            4: f"=DEGREES(ACOS(MAX(-1,MIN(1,-TAN(RADIANS({model}$B$17))*TAN(RADIANS(C{r}))))))",
            5: f"=SUMIFS($N$21:$N$1172,$A$21:$A$1172,A{r})",
            6: f"=SUMIFS($O$21:$O$1172,$A$21:$A$1172,A{r})",
            7: f"=IF(E{r}=0,0,F{r}/E{r})",
            8: f"=(24*60/PI())*{reg}$H$5*(1+0.033*COS(RADIANS(360*B{r}/365)))"
               f"*(COS(RADIANS({model}$B$17))*COS(RADIANS(C{r}))*SIN(RADIANS(D{r}))"
               f"+RADIANS(D{r})*SIN(RADIANS({model}$B$17))*SIN(RADIANS(C{r})))/3.6",
        }
        fmts = {1: "0", 2: "0", 3: NUM_2, 4: NUM_2, 5: NUM_2, 6: NUM_2, 7: NUM_3, 8: NUM_2}
        for col in range(1, 9):
            c = ws.cell(row=r, column=col, value=vals[col])
            c.font = F_BODY
            c.alignment = AL_C
            c.border = BOX
            c.number_format = fmts[col]
            if m % 2 == 0:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 17

    ws.cell(row=3, column=10, value="적용 입력값").font = F_SECTION
    ws.cell(row=3, column=10).fill = FILL_SECTION
    ws.merge_cells("J3:L3")
    for col in (11, 12):
        ws.cell(row=3, column=col).fill = FILL_SECTION
    for i, (label, formula, unit) in enumerate([
        ("위도 φ", f"={model}B17", "°N"), ("경사각 β", f"={model}B7", "°"), ("방위각 A", f"={model}B8", "°"),
    ]):
        r = 4 + i
        ws.cell(row=r, column=10, value=label).font = F_BODY_B
        v = ws.cell(row=r, column=11, value=formula)
        v.font = F_LINK
        v.alignment = AL_C
        v.number_format = "0.0000"
        ws.cell(row=r, column=12, value=unit).font = F_NOTE
        for col in (10, 11, 12):
            ws.cell(row=r, column=col).border = BOX

    section(ws, 18, "■ 15분 간격 태양벡터 적분 상세 (근거 자료 · 행 그룹으로 접혀 있음)", "Y")
    ws.cell(row=19, column=16,
            value=f"↓ '{SCEN}' 시나리오별 cosAOI⁺ (시나리오 경사각·방위각 적용)").font = F_NOTE
    ws.merge_cells(start_row=19, start_column=16, end_row=19, end_column=25)
    header_row(ws, 20, ["월", "Step", "태양시\n[hour]", "DOY", "적위 δ", "시각각 ω", "Sun_E", "Sun_N",
                        "Sun_U=cosθz", "Normal_E", "Normal_N", "Normal_U", "cosAOI", "cosθz⁺", "cosAOI⁺"]
                       + [f"cosAOI⁺\nS{i + 1}" for i in range(len(SCENARIOS))], height=30)
    r = 21
    for m in range(1, 13):
        for step in range(1, STEPS + 1):
            vals = {
                1: m,
                2: step,
                3: f"=(B{r}-0.5)*{reg}$H$6",
                4: MONTH_DOY[m - 1],
                5: f"=INDEX($C$4:$C$15,MATCH(A{r},$A$4:$A$15,0))",
                6: f"=15*(C{r}-12)",
                7: f"=-COS(RADIANS(E{r}))*SIN(RADIANS(F{r}))",
                8: f"=COS(RADIANS({model}$B$17))*SIN(RADIANS(E{r}))"
                   f"-SIN(RADIANS({model}$B$17))*COS(RADIANS(E{r}))*COS(RADIANS(F{r}))",
                9: f"=SIN(RADIANS({model}$B$17))*SIN(RADIANS(E{r}))"
                   f"+COS(RADIANS({model}$B$17))*COS(RADIANS(E{r}))*COS(RADIANS(F{r}))",
                10: f"=SIN(RADIANS({model}$B$7))*SIN(RADIANS({model}$B$8))",
                11: f"=SIN(RADIANS({model}$B$7))*COS(RADIANS({model}$B$8))",
                12: f"=COS(RADIANS({model}$B$7))",
                13: f"=G{r}*J{r}+H{r}*K{r}+I{r}*L{r}",
                14: f"=MAX(0,I{r})",
                15: f"=IF(I{r}>0,MAX(0,M{r}),0)",
            }
            for s in range(len(SCENARIOS)):
                s_row = 5 + s
                b = f"{q(SCEN)}$B${s_row}"
                a = f"{q(SCEN)}$C${s_row}"
                vals[16 + s] = (
                    f"=IF($I{r}>0,MAX(0,$G{r}*SIN(RADIANS({b}))*SIN(RADIANS({a}))"
                    f"+$H{r}*SIN(RADIANS({b}))*COS(RADIANS({a}))+$I{r}*COS(RADIANS({b}))),0)"
                )
            for col in range(1, 16 + len(SCENARIOS)):
                c = ws.cell(row=r, column=col, value=vals[col])
                c.font = Font(name=FONT, size=8.5)
                c.alignment = AL_C
                c.number_format = "0" if col in (1, 2, 4) else "0.0000"
            r += 1
    ws.row_dimensions.group(21, r - 1, outline_level=1, hidden=True)

    widths(ws, {"A": 7, "B": 11, "C": 12, "D": 14, "E": 12, "F": 12, "G": 11, "H": 14, "I": 4, "J": 13,
                "K": 12, "L": 8, "M": 12, "N": 11, "O": 11})
    for col in range(16, 16 + len(SCENARIOS)):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = "A21"
    page_setup(ws)
    ws.sheet_properties.tabColor = BLUE
    return ws


# ==========================================================================
# 16_적용한계·참고
# ==========================================================================
INPUT_RULES = [
    ["경사각 β", "0° = 수평 / 90° = 수직", "고정식 국내 일반 20~30°"],
    ["방위각 A", "0° 북 / 90° 동 / 180° 남 / 270° 서", "정남향 = 180°"],
    ["지면반사율 ρg", "일반 토지·잔디 0.15~0.25", "적설 시 크게 상승 가능하므로 별도 검토"],
    ["PR", "온도·인버터·배선·오염·불일치·음영 등 포함 범위를 프로젝트마다 명시", "포함 범위를 문서화하지 않으면 이중 차감 위험"],
    ["가동률 / 출력제한", "PR과 중복되지 않도록 손실 경계를 명확히 구분", "고장·정비는 가동률, 의도적 감발은 출력제한"],
]

LIMIT_ROWS = [
    "본 모델은 광역 지역 월자료와 대표 위도를 사용하는 예비 타당성 분석용입니다.",
    "월평균 GHI만으로 DNI/DHI를 통계적으로 분리하므로, 시간별 실측 또는 위성자료 기반 모델보다 불확실성이 큽니다.",
    "지형·수평선·주변 음영, 적설, 모듈 온도, IAM, 양면발전, 배열 간 차폐, 인버터 클리핑은 별도 모델링하지 않습니다.",
    "금융·EPC 성능보증·정밀 설계에는 현장 좌표, 시간별 GHI/DNI/DHI 또는 POA, 모듈·인버터 사양과 손실 시뮬레이션을 사용하세요.",
    "현장 POA 실측이 있으면 경사면 전사 계산 대신 실측값을 우선 적용하는 것이 적절합니다.",
]

REFERENCES = [
    ["기상청 기상자료개방포털", "기후통계분석 - 일조일사 (원자료 출처)", "https://data.kma.go.kr/"],
    ["Erbs, Klein & Duffie (1982)", "월평균 산란비(diffuse fraction) 상관식",
     "https://doi.org/10.1016/0038-092X(82)90302-4"],
    ["Sandia PVPMC", "Plane of Array (POA) Irradiance",
     "https://pvpmc.sandia.gov/modeling-guide/1-weather-design-inputs/plane-of-array-poa-irradiance/"],
    ["Sandia PVPMC", "Calculating POA Irradiance (등방성 천공산란·지면반사·입사각)",
     "https://pvpmc.sandia.gov/modeling-guide/1-weather-design-inputs/plane-of-array-poa-irradiance/calculating-poa-irradiance"],
    ["Sandia PVPMC", "PV Performance Metrics - Performance Ratio",
     "https://pvpmc.sandia.gov/modeling-guide/5-ac-system-output/pv-performance-metrics/performance-ratio/"],
]


def build_limit(wb):
    ws = wb.create_sheet(LIMIT)
    title_block(ws, "적용 한계 및 참고 자료", "본 워크북의 계산 결과를 해석할 때 반드시 함께 확인해야 하는 입력 기준·한계·근거 자료입니다.", "D")
    section(ws, 4, "■ 입력 기준", "D")
    header_row(ws, 5, ["항목", "기준", "비고"], height=20)
    ws.merge_cells("C5:D5")
    ws.cell(row=5, column=4).fill = FILL_HEAD
    ws.cell(row=5, column=4).border = BOX
    for i, row in enumerate(INPUT_RULES):
        r = 6 + i
        for col, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = F_BODY_B if col == 1 else F_BODY
            c.alignment = AL_LW
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BOX
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 30

    start = 6 + len(INPUT_RULES) + 1
    section(ws, start, "■ 적용 한계 및 사용 권고", "D")
    for i, text in enumerate(LIMIT_ROWS):
        r = start + 1 + i
        ws.cell(row=r, column=1, value="•").alignment = AL_C
        ws.cell(row=r, column=1).font = F_BODY
        c = ws.cell(row=r, column=2, value=text)
        c.font = F_BODY
        c.alignment = AL_LW
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BOX
        ws.row_dimensions[r].height = 26

    start2 = start + 1 + len(LIMIT_ROWS) + 1
    section(ws, start2, "■ 참고 문헌 및 공식 자료", "D")
    header_row(ws, start2 + 1, ["출처", "내용", "URL"], height=20)
    ws.merge_cells(start_row=start2 + 1, start_column=3, end_row=start2 + 1, end_column=4)
    ws.cell(row=start2 + 1, column=4).fill = FILL_HEAD
    ws.cell(row=start2 + 1, column=4).border = BOX
    for i, (src, desc, url) in enumerate(REFERENCES):
        r = start2 + 2 + i
        ws.cell(row=r, column=1, value=src).font = F_BODY_B
        ws.cell(row=r, column=2, value=desc).font = F_BODY
        u = ws.cell(row=r, column=3, value=url)
        u.font = Font(name=FONT, size=8.5, color="0563C1", underline="single")
        u.hyperlink = url
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).alignment = AL_LW
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_BAND
        ws.row_dimensions[r].height = 26

    widths(ws, {"A": 22, "B": 46, "C": 46, "D": 22})
    page_setup(ws, landscape=False)
    ws.sheet_properties.tabColor = BLUE
    return ws


# ==========================================================================
# 20_원자료
# ==========================================================================
RAW_HEAD = ["지역명", "연도", "월", "연월", "일조합_hr", "일사합_kWh_m2", "일수", "일평균_일사합_kWh_m2"]


def read_source_raw(src_path: str):
    """원본 워크북의 '원자료' 시트에서 실측값(지역·연도·월·연월·일조합·일사합)을 읽는다."""
    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb["원자료"]
    rows = []
    for r in range(2, ws.max_row + 1):
        region = ws.cell(row=r, column=1).value
        if region is None:
            continue
        rows.append([
            region,
            ws.cell(row=r, column=2).value,
            ws.cell(row=r, column=3).value,
            ws.cell(row=r, column=4).value,
            ws.cell(row=r, column=5).value,
            ws.cell(row=r, column=6).value,
        ])
    wb.close()
    return rows


def build_raw(wb, raw_rows):
    ws = wb.create_sheet(RAW)
    for i, h in enumerate(RAW_HEAD):
        c = ws.cell(row=1, column=1 + i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_C
        c.border = BOX
    ws.row_dimensions[1].height = 24

    for i, row in enumerate(raw_rows):
        r = 2 + i
        for col, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = F_BODY
            c.alignment = AL_C
            c.number_format = NUM_1 if col in (5, 6) else "General"
        d = ws.cell(row=r, column=7, value=f"=DAY(EOMONTH(DATE(B{r},C{r},1),0))")
        d.font = F_BODY
        d.alignment = AL_C
        d.number_format = "0"
        a = ws.cell(row=r, column=8, value=f"=F{r}/G{r}")
        a.font = F_BODY
        a.alignment = AL_C
        a.number_format = NUM_2

    last = 1 + len(raw_rows)
    table = Table(displayName="RawDataKWh", ref=f"A1:H{last}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(table)

    widths(ws, {"A": 12, "B": 9, "C": 7, "D": 11, "E": 13, "F": 17, "G": 8, "H": 22})
    ws.freeze_panes = "A2"
    page_setup(ws, landscape=False, rows_to_repeat="1:1")
    ws.sheet_properties.tabColor = "548235"
    return ws


# ==========================================================================
# 21~23 피벗 (월 합계 / 일평균을 한 시트에 통합)
# ==========================================================================
def _pivot_header(ws, title, subtitle, left_label, right_label, key_head):
    title_block(ws, title, subtitle, "Y")
    heads = [f"{key_head}\n({left_label})"] + REGIONS + ["전국평균"]
    heads_r = [f"{key_head}\n({right_label})"] + REGIONS + ["전국평균"]
    for i, h in enumerate(heads):
        c = ws.cell(row=3, column=1 + i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = AL_CW
        c.border = BOX
        c2 = ws.cell(row=3, column=14 + i, value=heads_r[i])
        c2.font = F_HEAD
        c2.fill = FILL_HEAD
        c2.alignment = AL_CW
        c2.border = BOX
    ws.row_dimensions[3].height = 30


def _pivot_style(ws, first_row, last_row, fmt_left, fmt_right):
    for r in range(first_row, last_row + 1):
        for col in range(1, 13):
            c = ws.cell(row=r, column=col)
            c.font = F_BODY
            c.alignment = AL_C
            c.border = BOX
            if col > 1:
                c.number_format = fmt_left
            if (r - first_row) % 2 == 1:
                c.fill = FILL_BAND
        for col in range(14, 26):
            c = ws.cell(row=r, column=col)
            c.font = F_BODY
            c.alignment = AL_C
            c.border = BOX
            if col > 14:
                c.number_format = fmt_right
            if (r - first_row) % 2 == 1:
                c.fill = FILL_BAND
        ws.row_dimensions[r].height = 16
    widths(ws, {"A": 11, "M": 3, "N": 11})
    for col in list(range(2, 13)) + list(range(15, 26)):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = "A4"
    page_setup(ws, rows_to_repeat="3:3")
    ws.sheet_properties.tabColor = "548235"


def build_pivot_month(wb):
    ws = wb.create_sheet(PM)
    _pivot_header(ws, "연월 × 지역 일사량 피벗",
                  "왼쪽 표: 월 합계 [kWh/m²]  ·  오른쪽 표(N열~): 일평균 [kWh/m²/일]  ·  모두 20_원자료 표(RawDataKWh)에서 집계",
                  "월 합계", "일평균", "연월")
    ym = [f"{y}-{m:02d}" for y in YEARS for m in range(1, 13)]
    for i, key in enumerate(ym):
        r = 4 + i
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=14, value=f"=$A{r}")
        for j in range(10):
            col = 2 + j
            ws.cell(row=r, column=col,
                    value=f"=SUMIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[연월],$A{r},"
                          f"RawDataKWh[지역명],{get_column_letter(col)}$3)")
            col2 = 15 + j
            ws.cell(row=r, column=col2,
                    value=f"=AVERAGEIFS(RawDataKWh[일평균_일사합_kWh_m2],RawDataKWh[연월],$A{r},"
                          f"RawDataKWh[지역명],{get_column_letter(col2)}$3)")
        ws.cell(row=r, column=12, value=f"=AVERAGE(B{r}:K{r})")
        ws.cell(row=r, column=25, value=f"=AVERAGE(O{r}:X{r})")
    _pivot_style(ws, 4, 3 + len(ym), NUM_1, NUM_2)
    return ws


def build_pivot_year(wb):
    ws = wb.create_sheet(PY)
    _pivot_header(ws, "연도 × 지역 일사량 피벗",
                  "왼쪽 표: 연 합계 [kWh/m²/년]  ·  오른쪽 표(N열~): 일평균 [kWh/m²/일, 일수 가중]  ·  14행은 10개년 평균",
                  "연 합계", "일평균", "연도")
    for i, y in enumerate(YEARS):
        r = 4 + i
        ws.cell(row=r, column=1, value=y)
        ws.cell(row=r, column=14, value=f"=$A{r}")
        for j in range(10):
            col = 2 + j
            ws.cell(row=r, column=col,
                    value=f"=SUMIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[연도],$A{r},"
                          f"RawDataKWh[지역명],{get_column_letter(col)}$3)")
            col2 = 15 + j
            letter = get_column_letter(col2)
            ws.cell(row=r, column=col2,
                    value=f"=SUMIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[연도],$A{r},RawDataKWh[지역명],{letter}$3)"
                          f"/SUMIFS(RawDataKWh[일수],RawDataKWh[연도],$A{r},RawDataKWh[지역명],{letter}$3)")
        ws.cell(row=r, column=12, value=f"=AVERAGE(B{r}:K{r})")
        ws.cell(row=r, column=25, value=f"=AVERAGE(O{r}:X{r})")
    _pivot_style(ws, 4, 13, NUM_1, NUM_2)

    ws.cell(row=14, column=1, value="10개년 평균")
    ws.cell(row=14, column=14, value="10개년 평균")
    for col in list(range(2, 13)) + list(range(15, 26)):
        letter = get_column_letter(col)
        ws.cell(row=14, column=col, value=f"=AVERAGE({letter}4:{letter}13)")
    for col in list(range(1, 13)) + list(range(14, 26)):
        c = ws.cell(row=14, column=col)
        c.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = AL_C
        c.border = BOX
        if col not in (1, 14):
            c.number_format = NUM_1 if col < 13 else NUM_2
    ws.row_dimensions[14].height = 18
    ws.cell(row=16, column=1,
            value="※ 일평균(오른쪽 표) = 해당 연도 일사합 총합 ÷ 해당 연도 총 일수 (일수 가중평균)").font = F_NOTE
    ws.merge_cells("A16:L16")
    return ws


def build_pivot_season(wb):
    ws = wb.create_sheet(PS)
    _pivot_header(ws, "달력월 기후평균 (계절 패턴)",
                  "왼쪽 표: 2016~2025 동일 월의 월 합계 평균 [kWh/m²]  ·  오른쪽 표(N열~): 동일 월의 일평균 [kWh/m²/일]",
                  "월 합계 평균", "일평균", "월")
    for m in range(1, 13):
        r = 3 + m
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=14, value=f"=$A{r}")
        for j in range(10):
            col = 2 + j
            ws.cell(row=r, column=col,
                    value=f"=AVERAGEIFS(RawDataKWh[일사합_kWh_m2],RawDataKWh[월],$A{r},"
                          f"RawDataKWh[지역명],{get_column_letter(col)}$3)")
            col2 = 15 + j
            ws.cell(row=r, column=col2,
                    value=f"=AVERAGEIFS(RawDataKWh[일평균_일사합_kWh_m2],RawDataKWh[월],$A{r},"
                          f"RawDataKWh[지역명],{get_column_letter(col2)}$3)")
        ws.cell(row=r, column=12, value=f"=AVERAGE(B{r}:K{r})")
        ws.cell(row=r, column=25, value=f"=AVERAGE(O{r}:X{r})")
    _pivot_style(ws, 4, 15, NUM_1, NUM_2)
    for r in range(4, 16):
        ws.cell(row=r, column=1).number_format = "0"
        ws.cell(row=r, column=14).number_format = "0"
    ws.cell(row=17, column=1,
            value="※ 2016~2025년(10개년) 동일 달력월의 평균값입니다. 단위 환산: 1 kWh/m² = 3.6 MJ/m²").font = F_NOTE
    ws.merge_cells("A17:L17")
    return ws


# ==========================================================================
# 24_차트
# ==========================================================================
def build_charts(wb):
    ws = wb.create_sheet(CHART)
    ws.sheet_view.showGridLines = False
    title_block(ws, "지역별 일사량 차트", "연도별 추이와 달력월 계절 패턴을 kWh/m² 단위로 표시합니다. 원본 데이터는 21~23 피벗 시트입니다.", "P")

    py_ws, ps_ws = wb[PY], wb[PS]
    specs = [
        (py_ws, "연도별 지역 일사량 추이 (kWh/m²/년)", "kWh/m²/년", 2, 11, 4, 13, 1, "A4"),
        (ps_ws, "달력월 계절 패턴 (kWh/m²/월)", "kWh/m²/월", 2, 11, 4, 15, 1, "A24"),
        (py_ws, "연도별 지역 일평균 일사량 (kWh/m²/일)", "kWh/m²/일", 15, 24, 4, 13, 14, "A44"),
        (ps_ws, "달력월 계절 패턴 - 일평균 (kWh/m²/일)", "kWh/m²/일", 15, 24, 4, 15, 14, "A64"),
    ]
    for src, title, y_title, min_col, max_col, min_row, max_row, cat_col, anchor in specs:
        ch = LineChart()
        ch.title = title
        ch.style = 2
        ch.height, ch.width = 9.5, 26
        ch.y_axis.title = y_title
        data = Reference(src, min_col=min_col, max_col=max_col, min_row=3, max_row=max_row)
        cats = Reference(src, min_col=cat_col, min_row=min_row, max_row=max_row)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        for s in ch.series:
            s.smooth = False
        ws.add_chart(ch, anchor)

    ws.column_dimensions["A"].width = 14
    ws.sheet_properties.tabColor = "548235"
    return ws


# ==========================================================================
# main
# ==========================================================================
def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 1
    src_path, out_path = sys.argv[1], sys.argv[2]
    raw_rows = read_source_raw(src_path)
    print(f"원자료 {len(raw_rows):,}행 로드")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_version(wb)
    build_overview(wb)
    build_calc(wb)
    build_factor(wb)
    build_diag(wb)
    build_dict(wb)
    build_test(wb)
    build_report(wb)
    build_dev(wb)
    build_model(wb)
    build_mfact(wb)
    build_scenario(wb)
    build_region(wb)
    build_geo(wb)
    build_limit(wb)
    build_raw(wb, raw_rows)
    build_pivot_month(wb)
    build_pivot_year(wb)
    build_pivot_season(wb)
    build_charts(wb)
    build_scenario_helper(wb)

    for ws in wb.worksheets:
        if ws.sheet_state == "visible" and not ws.sheet_properties.tabColor:
            ws.sheet_properties.tabColor = NAVY
        ws.sheet_view.zoomScale = 100

    wb.properties.title = "태양광 발전량 성능진단 기준모델"
    wb.properties.subject = f"문서 {DOC_VERSION} / 모델 {MODEL_VERSION}"
    wb.properties.creator = "PrimeGrid"
    wb.active = 0
    wb.save(out_path)
    print(f"저장 완료: {out_path} ({len(wb.worksheets)} 시트)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
